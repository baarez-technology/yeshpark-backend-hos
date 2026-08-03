"""
Channel Manager API Endpoints
Provides comprehensive channel manager functionality for OTA connections, room mappings,
rate sync, restrictions, promotions, and sync logs.

This API supports the frontend Channel Manager interface and integrates with
the dummy channel manager service.
"""
import logging
import asyncio
import csv
import io
import hashlib
import re
import time
import uuid
from typing import List, Optional, Dict, Any, Union
from datetime import date, datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query, Body, Path
from fastapi.responses import Response, StreamingResponse
from sqlmodel import select, and_, or_, func
from sqlmodel.ext.asyncio.session import AsyncSession
from pydantic import BaseModel, Field, field_validator, ConfigDict
from pydantic import AliasChoices

# PDF export support
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel export support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.db.session import get_tenant_session
from app.api.v1.auth import get_current_user, get_current_user_optional
from app.models.channel_manager import (
    OTAConnection, OTARoomMapping, OTARateMapping,
    AvailabilityGrid, ChannelRestriction, SyncLog
)
from app.models.inventory import RoomType, Room, RatePlan, DailyAvailability, DailyRate, PromoCode
from app.models.reservations import Booking, Guest
from app.models.reviews import Review
from app.models.user import User

router = APIRouter()
logger = logging.getLogger(__name__)

# In-memory cache for tracking export requests (prevents duplicate exports)
# Key: user_id + request_hash, Value: timestamp
_export_request_cache: Dict[str, float] = {}
_EXPORT_COOLDOWN_SECONDS = 5  # Prevent duplicate exports within 5 seconds

# ============== HELPER FUNCTIONS ==============

def get_ota_booking_source(ota_code: str) -> str:
    """
    Map OTA code to booking source.
    This mapping is used to filter bookings when calculating stats.
    
    Mapping:
    - DUMMY → CRS (DUMMY OTA bookings have source "CRS")
    - BOOKING → Booking.com
    - EXPEDIA → Expedia
    - etc.
    """
    ota_source_map = {
        "DUMMY": "CRS",  # DUMMY OTA bookings have source "CRS"
        "BOOKING": "Booking.com",
        "EXPEDIA": "Expedia",
        "AGODA": "Agoda",
        "AIRBNB": "airbnb",
        "MMT": "MakeMyTrip",
        "TRIP": "Trip.com",
        "GOOGLE": "Google Hotel Ads",
        "CRS": "Dummy Channel Manager"
    }
    return ota_source_map.get(ota_code.upper(), ota_code)

def filter_bookings_by_ota(bookings: List[Booking], ota_code: str) -> List[Booking]:
    """
    Filter bookings by OTA code using the booking source mapping.
    Checks both the mapped source and the OTA code itself for flexibility.
    """
    expected_source = get_ota_booking_source(ota_code)
    return [
        b for b in bookings 
        if b.booking_source and (
            expected_source.lower() in b.booking_source.lower() or
            ota_code.lower() in b.booking_source.lower() or
            b.booking_source.lower() == expected_source.lower()
        )
    ]


def _build_ota_item(ota: OTAConnection, all_bookings: List[Booking], all_reviews: List) -> Dict[str, Any]:
    """Build one OTA item dict with stats (same shape as GET /otas and GET /otas/{id})."""
    ota_bookings = filter_bookings_by_ota(all_bookings, ota.ota_code)
    ota_revenue = sum(b.total_price for b in ota_bookings if b.status != "cancelled")
    source_map = {
        "BOOKING": "booking_com",
        "EXPEDIA": "expedia",
        "AGODA": "agoda",
        "AIRBNB": "airbnb",
        "TRIPADVISOR": "tripadvisor",
        "DUMMY": "dummy"
    }
    review_source = source_map.get(ota.ota_code.upper(), ota.ota_code.lower())
    ota_reviews = [
        r for r in all_reviews
        if r.source and (
            r.source == review_source or
            review_source in r.source.lower() or
            ota.ota_code.lower() in r.source.lower() or
            ota.ota_name.lower() in r.source.lower()
        )
    ]
    avg_rating = sum(r.overall_rating for r in ota_reviews) / len(ota_reviews) if ota_reviews else 0.0
    return {
        "id": ota.id,
        "name": ota.ota_name,
        "code": ota.ota_code,
        "logo": ota.logo_url,
        "status": ota.connection_status,
        "lastSync": ota.last_sync_at.isoformat() if ota.last_sync_at else None,
        "nextSync": ota.next_sync_at.isoformat() if ota.next_sync_at else None,
        "errorMessage": ota.error_message,
        "credentials": {
            "username": ota.api_username or "",
            "apiKey": "***" if ota.api_key_encrypted else "",
            "hotelId": ota.hotel_id_on_ota or ""
        },
        "syncSettings": {
            "autoSync": ota.auto_sync_enabled,
            "syncInterval": ota.sync_interval_minutes,
            "syncRates": ota.sync_rates,
            "syncAvailability": ota.sync_availability,
            "syncRestrictions": ota.sync_restrictions
        },
        "stats": {
            "totalBookings": len([b for b in ota_bookings if b.status != "cancelled"]),
            "revenue": ota_revenue,
            "avgRating": round(avg_rating, 2),
            "commission": ota.commission_rate or 0.0
        },
        "color": ota.brand_color
    }


# ============== SCHEMAS ==============

class OTAConnectionResponse(BaseModel):
    id: int
    name: str
    code: str
    logo: Optional[str] = None
    status: str  # connected, disconnected, error, syncing
    lastSync: Optional[str] = None
    nextSync: Optional[str] = None
    errorMessage: Optional[str] = None
    credentials: Dict[str, Any]
    syncSettings: Dict[str, Any]
    stats: Dict[str, Any]
    color: Optional[str] = None

class OTACreateRequest(BaseModel):
    name: str
    code: str
    logo: Optional[str] = None
    credentials: Dict[str, Any] = {}
    syncSettings: Optional[Dict[str, Any]] = None
    commission: Optional[float] = None

class RoomMappingResponse(BaseModel):
    id: int
    pmsRoomType: str
    pmsRoomTypeId: int
    pmsRoomCode: Optional[str] = None
    basePrice: float
    inventory: int
    otaMappings: List[Dict[str, Any]]

def _slugify_ota_room_code(s: str) -> str:
    """Convert OTA room type name to a safe code (e.g. 'Deluxe King' -> 'deluxe_king')."""
    if not s or not s.strip():
        return "room"
    t = re.sub(r"[^a-zA-Z0-9\s\-]+", "", s.strip())
    t = re.sub(r"[\s\-]+", "_", t).strip("_").lower()
    return t if t else "room"


class RoomMappingCreateRequest(BaseModel):
    """Create room mapping. Accepts camelCase or snake_case. pmsRoomTypeId can be int (DB id) or str (slug); slug is resolved to id. otaRoomId optional (derived from otaRoomType). pmsRoomType optional (fetched from DB if not provided)."""
    model_config = ConfigDict(extra="ignore")  # Ignore extra fields from frontend to avoid 422

    pmsRoomTypeId: Union[int, str] = Field(..., validation_alias=AliasChoices("pmsRoomTypeId", "pms_room_type_id"))
    pmsRoomType: Optional[str] = Field(None, validation_alias=AliasChoices("pmsRoomType", "pms_room_type"))
    otaCode: str = Field(..., validation_alias=AliasChoices("otaCode", "ota_code"))
    otaRoomType: str = Field(..., validation_alias=AliasChoices("otaRoomType", "ota_room_type"))
    otaRoomId: Optional[str] = Field(None, validation_alias=AliasChoices("otaRoomId", "ota_room_id"))
    maxGuests: Optional[int] = Field(None, validation_alias=AliasChoices("maxGuests", "max_guests"))
    defaultRatePlan: Optional[str] = Field(None, validation_alias=AliasChoices("defaultRatePlan", "default_rate_plan"))

    @field_validator("otaCode", "otaRoomType", mode="before")
    @classmethod
    def coerce_str_fields(cls, v: Any) -> str:
        """Coerce to string (frontend may send numbers or null)."""
        if v is None:
            return ""
        return str(v).strip() if isinstance(v, str) else str(v)

    @field_validator("maxGuests", mode="before")
    @classmethod
    def coerce_max_guests(cls, v: Any) -> Optional[int]:
        """Coerce optional maxGuests from string to int."""
        if v is None or v == "":
            return None
        try:
            return int(v) if not isinstance(v, int) else v
        except (TypeError, ValueError):
            return None


class RoomMappingBulkItem(BaseModel):
    """Single item for bulk room mapping. Accepts camelCase or snake_case."""
    pmsRoomTypeId: int = Field(..., validation_alias=AliasChoices("pmsRoomTypeId", "pms_room_type_id"))
    otaRoomType: str = Field(..., validation_alias=AliasChoices("otaRoomType", "ota_room_type"))
    otaRoomId: Optional[str] = Field(None, validation_alias=AliasChoices("otaRoomId", "ota_room_id"))


class RoomMappingBulkRequest(BaseModel):
    """Bulk create/update room mappings for an OTA. Accepts camelCase or snake_case."""
    otaCode: str = Field(..., validation_alias=AliasChoices("otaCode", "ota_code"))
    mappings: List[RoomMappingBulkItem]


class AutoMapRequest(BaseModel):
    """Request model for auto-map endpoint - supports both snake_case and camelCase"""
    ota_code: Optional[str] = Field(None, alias="otaCode")
    otaCode: Optional[str] = None

    class Config:
        populate_by_name = True

class RateCalendarEntry(BaseModel):
    date: str
    roomType: str
    rates: Dict[str, float]
    otaRates: Dict[str, float]
    availability: int
    stopSell: bool
    cta: bool
    ctd: bool

class RestrictionCreateRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")  # Ignore extra fields from frontend to avoid 422
    roomType: Union[str, List[str]]  # Room type name, "ALL", or list of names
    otaCode: Union[str, List[str]]  # OTA code, "ALL", or list of codes
    dateRange: Dict[str, str]  # { start: "YYYY-MM-DD", end: "YYYY-MM-DD" }
    restriction: Dict[str, Any]  # { minStay, maxStay, cta, ctd, stopSell }
    reason: Optional[str] = None

class PromotionCreateRequest(BaseModel):
    name: str
    description: str
    discountType: str  # "percentage" | "fixed"
    discountValue: float
    validFrom: str
    validTo: str
    otaCodes: List[str]
    roomTypes: List[str]
    minStay: Optional[int] = None
    bookingWindow: Optional[Dict[str, str]] = None

class SyncLogResponse(BaseModel):
    id: int
    timestamp: str
    otaCode: str
    otaName: str
    action: str
    status: str
    message: str
    details: Optional[Dict[str, Any]] = None

class ChannelStatsResponse(BaseModel):
    connectedOTAs: int
    disconnectedOTAs: int
    errorOTAs: int
    totalBookings: int
    totalRevenue: float
    mappedRoomTypes: int
    totalRoomTypes: int
    activeRestrictions: int
    rateParityIssues: List[Dict[str, Any]]
    lastSync: Optional[str] = None
    revenueTrend: List[float]
    bookingsTrend: List[int]
    channelPerformance: List[Dict[str, Any]]
    avgCommission: float
    avgConversionRate: float
    revenueGrowth: str
    bookingsGrowth: str

# ============== OTA CONNECTIONS ENDPOINTS ==============

@router.get("/otas", tags=["Channel Manager"])
async def get_otas(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get all OTA connections including DUMMY channel manager"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /otas - User: {user_info}")
    
    stmt = select(OTAConnection).order_by(OTAConnection.ota_name)
    result = await session.execute(stmt)
    otas = result.scalars().all()
    
    # Get bookings to calculate stats
    bookings_stmt = select(Booking).where(Booking.booking_source.isnot(None))
    bookings_result = await session.execute(bookings_stmt)
    all_bookings = bookings_result.scalars().all()
    
    # Get reviews for rating calculation
    reviews_stmt = select(Review)
    reviews_result = await session.execute(reviews_stmt)
    all_reviews = reviews_result.scalars().all()
    
    items = []
    for ota in otas:
        items.append(_build_ota_item(ota, all_bookings, all_reviews))
    
    logger.info(f"[CHANNEL_MANAGER] Found {len(items)} OTA connections")
    return {"success": True, "data": {"items": items, "total": len(items)}}


# Known OTA platforms for "Add New Connection" – frontend can search/filter this list
AVAILABLE_OTA_PLATFORMS = [
    {"code": "BOOKING", "name": "Booking.com", "logo": "https://cf.bstatic.com/static/img/favicon/9ca83ba2a5a3293ff07452cb24949a5843af4592.svg", "brandColor": "#003580", "description": "World's leading accommodation provider"},
    {"code": "EXPEDIA", "name": "Expedia", "logo": "https://www.expedia.com/favicon.ico", "brandColor": "#FFD700", "description": "Full-service travel brand"},
    {"code": "AGODA", "name": "Agoda", "logo": "https://www.agoda.com/favicon.ico", "brandColor": "#E51937", "description": "Online travel platform"},
    {"code": "AIRBNB", "name": "Airbnb", "logo": "https://www.airbnb.com/favicon.ico", "brandColor": "#FF5A5F", "description": "Vacation rentals and experiences"},
    {"code": "HOTELS", "name": "Hotels.com", "logo": "https://www.hotels.com/favicon.ico", "brandColor": "#8B5CF6", "description": "Hotel booking platform"},
    {"code": "MMT", "name": "MakeMyTrip", "logo": "https://www.makemytrip.com/favicon.ico", "brandColor": "#F97316", "description": "Leading travel company in India"},
    {"code": "TRIPADVISOR", "name": "Tripadvisor", "logo": "https://static.tacdn.com/img/favicon.ico", "brandColor": "#34E0A1", "description": "Travel reviews and bookings"},
    {"code": "DUMMY", "name": "Dummy Channel Manager", "logo": None, "brandColor": "#A57865", "description": "Test integration for development"},
    {"code": "CUSTOM", "name": "Custom / Other OTA", "logo": None, "brandColor": "#6B7280", "description": "Add a custom OTA with your own name and code"},
]


@router.get("/platforms", tags=["Channel Manager"])
async def get_available_platforms(
    search: Optional[str] = Query(None, description="Search by name or code (case-insensitive)"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get available OTA platforms for Add New Connection. Frontend can search/filter; use CUSTOM to add or request new OTAs."""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /platforms - search={search}, User: {user_info}")

    # Get already connected OTA codes to filter them out
    existing_stmt = select(OTAConnection.ota_code)
    existing_result = await session.execute(existing_stmt)
    existing_codes = {row[0].upper() for row in existing_result.all()}

    # Filter out already connected platforms
    platforms = [p for p in AVAILABLE_OTA_PLATFORMS if p["code"].upper() not in existing_codes]

    if search and search.strip():
        q = search.strip().lower()
        platforms = [p for p in platforms if q in (p["name"] or "").lower() or q in (p["code"] or "").lower() or q in (p.get("description") or "").lower()]
    return {"success": True, "data": {"items": platforms, "total": len(platforms)}}


@router.get("/otas/{ota_id}", tags=["Channel Manager"])
async def get_ota(
    ota_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get specific OTA connection"""
    logger.info(f"[CHANNEL_MANAGER] GET /otas/{ota_id} - User: {current_user.email}")
    
    ota = await session.get(OTAConnection, ota_id)
    if not ota:
        logger.warning(f"[CHANNEL_MANAGER] OTA {ota_id} not found")
        raise HTTPException(status_code=404, detail="OTA connection not found")
    
    # Calculate stats from bookings
    bookings_stmt = select(Booking).where(Booking.booking_source.isnot(None))
    bookings_result = await session.execute(bookings_stmt)
    all_bookings = bookings_result.scalars().all()
    
    # Filter bookings for this OTA using helper function
    ota_bookings = filter_bookings_by_ota(all_bookings, ota.ota_code)
    bookings = [b for b in ota_bookings if b.status != "cancelled"]
    
    total_bookings = len(bookings)
    total_revenue = sum(b.total_price for b in bookings)
    
    # Get reviews for rating
    source_map = {
        "BOOKING": "booking_com",
        "EXPEDIA": "expedia",
        "AGODA": "agoda",
        "AIRBNB": "airbnb",
        "TRIPADVISOR": "tripadvisor",
        "DUMMY": "dummy"
    }
    review_source = source_map.get(ota.ota_code.upper(), ota.ota_code.lower())
    reviews_stmt = select(Review).where(
        or_(
            Review.source == review_source,
            Review.source.like(f"%{ota.ota_code.lower()}%"),
            Review.source.like(f"%{ota.ota_name.lower()}%")
        )
    )
    reviews_result = await session.execute(reviews_stmt)
    ota_reviews = reviews_result.scalars().all()
    avg_rating = sum(r.overall_rating for r in ota_reviews) / len(ota_reviews) if ota_reviews else 0.0
    
    return {
        "success": True,
        "data": {
            "id": ota.id,
            "name": ota.ota_name,
            "code": ota.ota_code,
            "logo": ota.logo_url,
            "status": ota.connection_status,
            "lastSync": ota.last_sync_at.isoformat() if ota.last_sync_at else None,
            "nextSync": ota.next_sync_at.isoformat() if ota.next_sync_at else None,
            "errorMessage": ota.error_message,
            "credentials": {
                "username": ota.api_username or "",
                "apiKey": "***",
                "hotelId": ota.hotel_id_on_ota or ""
            },
            "syncSettings": {
                "autoSync": ota.auto_sync_enabled,
                "syncInterval": ota.sync_interval_minutes,
                "syncRates": ota.sync_rates,
                "syncAvailability": ota.sync_availability,
                "syncRestrictions": ota.sync_restrictions
            },
            "stats": {
                "totalBookings": total_bookings,
                "revenue": total_revenue,
                "avgRating": round(avg_rating, 2),
                "commission": ota.commission_rate or 0.0
            },
            "color": ota.brand_color
        }
    }


@router.post("/otas", tags=["Channel Manager"])
async def create_ota(
    payload: OTACreateRequest,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Create/Connect new OTA (including DUMMY channel manager)"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] POST /otas - Code: {payload.code}, User: {user_info}")
    
    # Check if OTA already exists
    existing_stmt = select(OTAConnection).where(OTAConnection.ota_code == payload.code)
    existing_result = await session.execute(existing_stmt)
    existing = existing_result.scalar_one_or_none()

    if existing:
        # If already connected, return error
        if existing.connection_status == "connected" and existing.is_active:
            logger.warning(f"[CHANNEL_MANAGER] OTA with code {payload.code} already connected")
            raise HTTPException(status_code=400, detail=f"OTA with code {payload.code} is already connected")

        # Reconnect existing disconnected OTA
        logger.info(f"[CHANNEL_MANAGER] Reconnecting existing OTA: {payload.code}")
        existing.connection_status = "connected"
        existing.is_active = True
        existing.error_message = None
        existing.updated_at = datetime.utcnow()

        # Update credentials if provided
        credentials = payload.credentials or {}
        if credentials.get("username"):
            existing.api_username = credentials["username"]
        if credentials.get("apiKey"):
            existing.api_key_encrypted = credentials["apiKey"]
        if credentials.get("hotelId"):
            existing.hotel_id_on_ota = credentials["hotelId"]

        await session.commit()
        await session.refresh(existing)

        # Return full OTA with stats
        bookings_stmt = select(Booking).where(Booking.booking_source.isnot(None))
        bookings_result = await session.execute(bookings_stmt)
        all_bookings = bookings_result.scalars().all()
        reviews_stmt = select(Review)
        reviews_result = await session.execute(reviews_stmt)
        all_reviews = reviews_result.scalars().all()
        item = _build_ota_item(existing, all_bookings, all_reviews)

        logger.info(f"[CHANNEL_MANAGER] Reconnected OTA: {existing.ota_name}")
        return {"success": True, "data": item, "message": "OTA reconnected successfully"}

    # For DUMMY OTA, always succeed connection
    connection_status = "connected" if payload.code == "DUMMY" else "disconnected"

    # Use defaults if syncSettings not provided
    sync_settings = payload.syncSettings or {}
    credentials = payload.credentials or {}

    ota = OTAConnection(
        property_id=1,  # Default property
        ota_code=payload.code,
        ota_name=payload.name,
        logo_url=payload.logo,
        connection_status=connection_status,
        api_username=credentials.get("username"),
        api_key_encrypted=credentials.get("apiKey"),  # Should encrypt in production
        hotel_id_on_ota=credentials.get("hotelId"),
        auto_sync_enabled=sync_settings.get("autoSync", True),
        sync_interval_minutes=sync_settings.get("syncInterval", 5),
        sync_rates=sync_settings.get("syncRates", True),
        sync_availability=sync_settings.get("syncAvailability", True),
        sync_restrictions=sync_settings.get("syncRestrictions", True),
        commission_rate=payload.commission or 0.0,
        brand_color="#A57865" if payload.code == "DUMMY" else None,
        is_active=True
    )
    
    session.add(ota)
    await session.commit()
    await session.refresh(ota)
    
    logger.info(f"[CHANNEL_MANAGER] Created OTA connection: {ota.ota_name} (ID: {ota.id})")
    
    return {
        "success": True,
        "data": {
            "id": ota.id,
            "name": ota.ota_name,
            "code": ota.ota_code,
            "status": ota.connection_status
        }
    }


@router.put("/otas/{ota_id}", tags=["Channel Manager"])
async def update_ota(
    ota_id: int,
    payload: Dict[str, Any],
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Update OTA connection (credentials, settings)"""
    logger.info(f"[CHANNEL_MANAGER] PUT /otas/{ota_id} - User: {current_user.email}")
    
    ota = await session.get(OTAConnection, ota_id)
    if not ota:
        raise HTTPException(status_code=404, detail="OTA connection not found")
    
    # Update credentials if provided
    if "credentials" in payload:
        creds = payload["credentials"]
        if "username" in creds:
            ota.api_username = creds["username"]
        if "apiKey" in creds:
            ota.api_key_encrypted = creds["apiKey"]  # Should encrypt
        if "hotelId" in creds:
            ota.hotel_id_on_ota = creds["hotelId"]
    
    # Update sync settings if provided
    if "syncSettings" in payload:
        settings = payload["syncSettings"]
        if "autoSync" in settings:
            ota.auto_sync_enabled = settings["autoSync"]
        if "syncInterval" in settings:
            ota.sync_interval_minutes = settings["syncInterval"]
        if "syncRates" in settings:
            ota.sync_rates = settings["syncRates"]
        if "syncAvailability" in settings:
            ota.sync_availability = settings["syncAvailability"]
        if "syncRestrictions" in settings:
            ota.sync_restrictions = settings["syncRestrictions"]
    
    # Reconnect: when connection_status is set to "connected", restore is_active so data/sync work
    if "connection_status" in payload or "status" in payload:
        new_status = payload.get("connection_status") or payload.get("status")
        if new_status == "connected":
            ota.connection_status = "connected"
            ota.is_active = True
            ota.error_message = None
        elif new_status == "disconnected":
            ota.connection_status = "disconnected"
            ota.is_active = False
    
    ota.updated_at = datetime.utcnow()
    await session.commit()
    await session.refresh(ota)
    
    # Return full OTA with stats so frontend shows correct data after reconnect (no stale/missing stats)
    bookings_stmt = select(Booking).where(Booking.booking_source.isnot(None))
    bookings_result = await session.execute(bookings_stmt)
    all_bookings = bookings_result.scalars().all()
    reviews_stmt = select(Review)
    reviews_result = await session.execute(reviews_stmt)
    all_reviews = reviews_result.scalars().all()
    item = _build_ota_item(ota, all_bookings, all_reviews)
    
    logger.info(f"[CHANNEL_MANAGER] Updated OTA connection: {ota.ota_name}")
    
    return {"success": True, "message": "OTA connection updated successfully", "data": item}


@router.delete("/otas/{ota_id}", tags=["Channel Manager"])
async def delete_ota(
    ota_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Disconnect OTA"""
    logger.info(f"[CHANNEL_MANAGER] DELETE /otas/{ota_id} - User: {current_user.email}")
    
    ota = await session.get(OTAConnection, ota_id)
    if not ota:
        raise HTTPException(status_code=404, detail="OTA connection not found")
    
    ota.connection_status = "disconnected"
    ota.is_active = False
    ota.updated_at = datetime.utcnow()
    await session.commit()
    
    logger.info(f"[CHANNEL_MANAGER] Disconnected OTA: {ota.ota_name}")
    
    return {"success": True, "message": "OTA disconnected successfully"}


@router.post("/otas/{ota_id}/test", tags=["Channel Manager"])
async def test_ota_connection(
    ota_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Test OTA connection"""
    logger.info(f"[CHANNEL_MANAGER] POST /otas/{ota_id}/test - User: {current_user.email}")
    
    ota = await session.get(OTAConnection, ota_id)
    if not ota:
        raise HTTPException(status_code=404, detail="OTA connection not found")
    
    # For DUMMY OTA, always succeed
    if ota.ota_code == "DUMMY":
        logger.info(f"[CHANNEL_MANAGER] DUMMY OTA connection test - always succeeds")
        return {
            "success": True,
            "data": {
                "connected": True,
                "message": "Dummy Channel Manager connection successful",
                "responseTime": 50
            }
        }
    
    # For other OTAs, would test actual API connection
    # For now, simulate
    return {
        "success": True,
        "data": {
            "connected": True,
            "message": "Connection test successful",
            "responseTime": 100
        }
    }


@router.post("/otas/test-connection", tags=["Channel Manager"])
async def test_ota_connection_pre(
    payload: Dict[str, Any] = Body(...),
    current_user: User = Depends(get_current_user)
):
    """
    Pre-connection credential validation.
    Validates credentials format before an OTA is actually created.
    """
    ota_code = payload.get("otaCode", "")
    credentials = payload.get("credentials", {})

    logger.info(f"[CHANNEL_MANAGER] POST /otas/test-connection - OTA: {ota_code}, User: {current_user.email}")

    if not ota_code:
        return {"success": False, "message": "OTA code is required."}

    username = credentials.get("username", "")
    api_key = credentials.get("apiKey", "") or credentials.get("api_key", "")
    hotel_id = credentials.get("hotelId", "") or credentials.get("hotel_id", "")

    if not username or not api_key or not hotel_id:
        return {"success": False, "message": "All credential fields are required (username, apiKey, hotelId)."}

    # For real OTAs, this would make an API call to validate.
    # For now, validate format and return success.
    return {
        "success": True,
        "message": "Credentials validated. Connection will be verified on first sync."
    }


@router.post("/otas/{ota_id}/sync", tags=["Channel Manager"])
async def trigger_ota_sync(
    ota_id: int,
    sync_type: str = Query("all", alias="syncType", description="Sync type: rates, availability, restrictions, all"),
    date_range: Optional[Dict[str, str]] = Body(None, alias="dateRange"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Trigger manual sync for specific OTA"""
    logger.info(f"[CHANNEL_MANAGER] POST /otas/{ota_id}/sync - Type: {sync_type}, User: {current_user.email}")
    
    ota = await session.get(OTAConnection, ota_id)
    if not ota:
        raise HTTPException(status_code=404, detail="OTA connection not found")
    
    if ota.connection_status != "connected":
        raise HTTPException(status_code=400, detail="OTA is not connected")
    
    # Create sync log
    sync_log = SyncLog(
        property_id=ota.property_id,
        ota_connection_id=ota_id,
        sync_type=sync_type,
        sync_direction="push",
        status="in_progress",
        started_at=datetime.utcnow()
    )
    session.add(sync_log)
    await session.commit()
    await session.refresh(sync_log)
    
    logger.info(f"[CHANNEL_MANAGER] Created sync log: ID={sync_log.id}, Type={sync_type}, StartedAt={sync_log.started_at}")
    
    # For DUMMY OTA, this would trigger sync in dummy channel manager
    # The dummy channel manager would then call Glimmora backend APIs
    
    return {
        "success": True,
        "data": {
            "syncId": sync_log.id,
            "startedAt": sync_log.started_at.isoformat() if sync_log.started_at else None,
            "status": "in_progress",
            "estimatedDuration": 30
        }
    }


@router.post("/otas/sync/all", tags=["Channel Manager"])
async def sync_all_otas(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Trigger sync for all connected OTAs"""
    logger.info(f"[CHANNEL_MANAGER] POST /otas/sync/all - User: {current_user.email}")
    
    stmt = select(OTAConnection).where(
        and_(
            OTAConnection.connection_status == "connected",
            OTAConnection.is_active == True
        )
    )
    result = await session.execute(stmt)
    connected_otas = result.scalars().all()
    
    created_logs = []
    for ota in connected_otas:
        sync_log = SyncLog(
            property_id=ota.property_id,
            ota_connection_id=ota.id,
            sync_type="full",
            sync_direction="push",
            status="pending",
            started_at=datetime.utcnow()
        )
        session.add(sync_log)
        created_logs.append(sync_log)
    
    await session.commit()
    for log in created_logs:
        await session.refresh(log)
    sync_ids = [log.id for log in created_logs if log.id is not None]
    
    logger.info(f"[CHANNEL_MANAGER] Triggered sync for {len(connected_otas)} OTAs, sync log IDs: {sync_ids}")
    
    return {
        "success": True,
        "data": {
            "syncIds": sync_ids,
            "status": "pending"
        }
    }

# ============== ROOM MAPPINGS ENDPOINTS ==============

@router.get("/room-mappings", tags=["Channel Manager"])
async def get_room_mappings(
    ota_code: Optional[str] = Query(None, alias="otaCode"),
    pms_room_type_id: Optional[int] = Query(None, alias="pmsRoomTypeId"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get all room mappings"""
    logger.info(f"[CHANNEL_MANAGER] GET /room-mappings - OTA: {ota_code}, RoomType: {pms_room_type_id}")
    
    # Get all room types
    room_types_stmt = select(RoomType)
    if pms_room_type_id:
        room_types_stmt = room_types_stmt.where(RoomType.id == pms_room_type_id)
    room_types_result = await session.execute(room_types_stmt)
    room_types = room_types_result.scalars().all()
    
    items = []
    for rt in room_types:
        # Get OTA mappings for this room type
        mappings_stmt = select(OTARoomMapping).where(OTARoomMapping.room_type_id == rt.id)
        if ota_code:
            # Get OTA connection ID
            ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
            ota_result = await session.execute(ota_stmt)
            ota = ota_result.scalar_one_or_none()
            if ota:
                mappings_stmt = mappings_stmt.where(OTARoomMapping.ota_connection_id == ota.id)
        
        mappings_result = await session.execute(mappings_stmt)
        ota_mappings = mappings_result.scalars().all()
        
        ota_mappings_list = []
        for mapping in ota_mappings:
            ota_conn = await session.get(OTAConnection, mapping.ota_connection_id)
            ota_mappings_list.append({
                "id": mapping.id,
                "otaCode": ota_conn.ota_code if ota_conn else "",
                "otaRoomType": mapping.ota_room_name or mapping.ota_room_code,
                "otaRoomId": mapping.ota_room_code,
                "otaRoomCode": mapping.ota_room_code,
                "status": "active" if mapping.is_active else "inactive",
                "lastSync": mapping.last_synced_at.isoformat() if mapping.last_synced_at else None
            })
        
        # Get total room count for this room type
        from app.models.inventory import Room
        room_count_stmt = select(func.count(Room.id)).where(Room.room_type_id == rt.id)
        room_count_result = await session.execute(room_count_stmt)
        total_rooms = room_count_result.scalar() or 0
        
        items.append({
            "id": rt.id,
            "pmsRoomType": rt.name,
            "pmsRoomTypeId": rt.id,
            "pmsRoomCode": rt.slug,
            "basePrice": rt.base_price,
            "inventory": total_rooms,
            "otaMappings": ota_mappings_list
        })
    
    logger.info(f"[CHANNEL_MANAGER] Found {len(items)} room mappings")
    return {"success": True, "data": {"items": items, "total": len(items)}}


@router.post("/room-mappings", tags=["Channel Manager"])
async def create_room_mapping(
    payload: RoomMappingCreateRequest,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Create new room mapping. otaRoomId optional (derived from otaRoomType when missing). pmsRoomTypeId can be int or slug (resolved to id)."""
    try:
        logger.info(f"[CHANNEL_MANAGER] POST /room-mappings - OTA: {payload.otaCode}, RoomType: {payload.pmsRoomTypeId}")
        print(f"[CHANNEL_MANAGER] Creating room mapping - OTA: {payload.otaCode}, RoomType: {payload.pmsRoomTypeId}")
        
        # Validate required fields
        if not payload.otaCode:
            raise HTTPException(status_code=400, detail="Missing required field: otaCode")
        if payload.pmsRoomTypeId is None or payload.pmsRoomTypeId == "":
            raise HTTPException(status_code=400, detail="Missing required field: pmsRoomTypeId")
        if not (payload.otaRoomType or "").strip():
            raise HTTPException(status_code=400, detail="Missing required field: otaRoomType (OTA room type name)")
        ota_room_id = (payload.otaRoomId or "").strip()
        if not ota_room_id:
            ota_room_id = _slugify_ota_room_code(payload.otaRoomType)
        
        # Resolve pmsRoomTypeId: accept int (DB id), numeric string, or slug string
        pms_room_type_id: int
        if isinstance(payload.pmsRoomTypeId, int):
            pms_room_type_id = payload.pmsRoomTypeId
        else:
            value = (str(payload.pmsRoomTypeId) or "").strip()
            if not value:
                raise HTTPException(status_code=400, detail="Missing required field: pmsRoomTypeId (or provide a room type slug)")
            # Try to parse as integer first (numeric string like "1", "2")
            try:
                pms_room_type_id = int(value)
            except ValueError:
                # Not a number, treat as slug
                slug_stmt = select(RoomType).where(RoomType.slug == value)
                slug_result = await session.execute(slug_stmt)
                room_by_slug = slug_result.scalar_one_or_none()
                if not room_by_slug:
                    raise HTTPException(status_code=404, detail=f"Room type with slug '{value}' not found")
                pms_room_type_id = room_by_slug.id
        
        # Get OTA connection
        ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == payload.otaCode)
        ota_result = await session.execute(ota_stmt)
        ota = ota_result.scalar_one_or_none()
        
        if not ota:
            error_msg = f"OTA with code {payload.otaCode} not found"
            logger.error(f"[CHANNEL_MANAGER] {error_msg}")
            raise HTTPException(status_code=404, detail=error_msg)
        
        # Verify room type exists and get its name
        room_type = await session.get(RoomType, pms_room_type_id)
        if not room_type:
            error_msg = f"Room type with ID {pms_room_type_id} not found"
            logger.error(f"[CHANNEL_MANAGER] {error_msg}")
            raise HTTPException(status_code=404, detail=error_msg)
        
        # Use room type name from payload if provided, otherwise use from database
        pms_room_type_name = payload.pmsRoomType or room_type.name

        # Check if mapping already exists
        existing_stmt = select(OTARoomMapping).where(
            and_(
                OTARoomMapping.ota_connection_id == ota.id,
                OTARoomMapping.room_type_id == pms_room_type_id
            )
        )
        existing_result = await session.execute(existing_stmt)
        existing = existing_result.scalar_one_or_none()
        
        if existing:
            # Update existing mapping
            existing.ota_room_code = ota_room_id
            existing.ota_room_name = payload.otaRoomType or pms_room_type_name
            existing.is_active = True
            existing.last_synced_at = datetime.utcnow()
            await session.commit()
            await session.refresh(existing)
            
            logger.info(f"[CHANNEL_MANAGER] Updated existing room mapping: {existing.id}")
            print(f"[CHANNEL_MANAGER] Updated existing room mapping: {existing.id}")
            return {
                "success": True,
                "data": {
                    "id": existing.id,
                    "message": "Mapping updated successfully",
                    "otaCode": payload.otaCode,
                    "pmsRoomTypeId": pms_room_type_id,
                    "otaRoomId": ota_room_id
                }
            }
        
        # Create new mapping
        mapping = OTARoomMapping(
            property_id=ota.property_id,
            ota_connection_id=ota.id,
            room_type_id=pms_room_type_id,
            ota_room_code=ota_room_id,
            ota_room_name=payload.otaRoomType or pms_room_type_name,
            is_active=True,
            sync_status="synced",
            last_synced_at=datetime.utcnow()
        )
        
        session.add(mapping)
        await session.commit()
        await session.refresh(mapping)
        
        logger.info(f"[CHANNEL_MANAGER] Created room mapping: ID={mapping.id}")
        print(f"[CHANNEL_MANAGER] Created room mapping: ID={mapping.id}")
        
        return {
            "success": True,
            "data": {
                "id": mapping.id,
                "message": "Mapping created successfully",
                "otaCode": payload.otaCode,
                "pmsRoomTypeId": pms_room_type_id,
                "otaRoomId": ota_room_id
            }
        }
    
    except HTTPException:
        raise
    except Exception as e:
        error_msg = f"Unexpected error creating room mapping: {str(e)}"
        logger.error(f"[CHANNEL_MANAGER] {error_msg}", exc_info=True)
        raise HTTPException(status_code=500, detail=error_msg)


@router.get("/room-mappings/bulk-view", tags=["Channel Manager"])
async def get_room_mappings_bulk_view(
    ota_code: str = Query(..., alias="otaCode", description="OTA code (e.g. BOOKING, DUMMY). Required to fetch bulk-view for one OTA."),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get all hotel room types and their mappings for an OTA in one response. Use this for a unified bulk-mapping UI: display all PMS room types with existing OTA mappings and known OTA room type suggestions."""
    logger.info(f"[CHANNEL_MANAGER] GET /room-mappings/bulk-view - OTA: {ota_code}")

    ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
    ota_result = await session.execute(ota_stmt)
    ota = ota_result.scalar_one_or_none()
    if not ota:
        raise HTTPException(status_code=404, detail=f"OTA with code {ota_code} not found")

    room_types_stmt = select(RoomType).where(RoomType.is_active == True).order_by(RoomType.name)
    room_types_result = await session.execute(room_types_stmt)
    room_types = room_types_result.scalars().all()

    mappings_stmt = select(OTARoomMapping).where(
        OTARoomMapping.ota_connection_id == ota.id,
        OTARoomMapping.is_active == True
    )
    mappings_result = await session.execute(mappings_stmt)
    mappings_list = mappings_result.scalars().all()
    mappings_by_room = {m.room_type_id: m for m in mappings_list}

    known_ota_room_types: List[Dict[str, str]] = []
    seen = set()
    for m in mappings_list:
        key = (m.ota_room_name or m.ota_room_code, m.ota_room_code)
        if key not in seen:
            seen.add(key)
            known_ota_room_types.append({"otaRoomType": m.ota_room_name or m.ota_room_code, "otaRoomId": m.ota_room_code})
    # When OTA has no mappings (e.g. Trip.com), dropdown would be empty. Add PMS room types as fallback options.
    for rt in room_types:
        ota_room_type = rt.name
        ota_room_id = (rt.slug or _slugify_ota_room_code(rt.name)) or "room"
        key = (ota_room_type, ota_room_id)
        if key not in seen:
            seen.add(key)
            known_ota_room_types.append({"otaRoomType": ota_room_type, "otaRoomId": ota_room_id})

    items: List[Dict[str, Any]] = []
    for rt in room_types:
        m = mappings_by_room.get(rt.id)
        room_count_stmt = select(func.count(Room.id)).where(Room.room_type_id == rt.id)
        room_count_result = await session.execute(room_count_stmt)
        total_rooms = room_count_result.scalar() or 0
        items.append({
            "pmsRoomTypeId": rt.id,
            "pmsRoomType": rt.name,
            "pmsRoomCode": rt.slug or "",
            "basePrice": float(rt.base_price or 0),
            "inventory": total_rooms,
            "mappingId": m.id if m else None,
            "otaRoomType": (m.ota_room_name or m.ota_room_code) if m else None,
            "otaRoomId": m.ota_room_code if m else None,
        })

    return {
        "success": True,
        "data": {
            "otaCode": ota_code,
            "items": items,
            "knownOtaRoomTypes": known_ota_room_types,
            "total": len(items),
        },
    }


@router.post("/room-mappings/bulk", tags=["Channel Manager"])
async def create_or_update_room_mappings_bulk(
    payload: RoomMappingBulkRequest,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Create or update multiple room mappings for an OTA in a single request. Supports bulk mapping from a unified screen with multi-select."""
    try:
        ota_code = payload.otaCode
        if not ota_code:
            raise HTTPException(status_code=400, detail="Missing required field: otaCode")

        ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
        ota_result = await session.execute(ota_stmt)
        ota = ota_result.scalar_one_or_none()
        if not ota:
            raise HTTPException(status_code=404, detail=f"OTA with code {ota_code} not found")

        created = 0
        updated = 0
        errors: List[Dict[str, Any]] = []

        for entry in payload.mappings:
            try:
                if not (entry.otaRoomType or "").strip():
                    errors.append({"pmsRoomTypeId": entry.pmsRoomTypeId, "error": "otaRoomType is required"})
                    continue
                ota_room_id = (entry.otaRoomId or "").strip() or _slugify_ota_room_code(entry.otaRoomType)

                room_type = await session.get(RoomType, entry.pmsRoomTypeId)
                if not room_type:
                    errors.append({"pmsRoomTypeId": entry.pmsRoomTypeId, "error": f"Room type ID {entry.pmsRoomTypeId} not found"})
                    continue

                existing_stmt = select(OTARoomMapping).where(
                    and_(
                        OTARoomMapping.ota_connection_id == ota.id,
                        OTARoomMapping.room_type_id == entry.pmsRoomTypeId
                    )
                )
                existing_result = await session.execute(existing_stmt)
                existing = existing_result.scalar_one_or_none()

                if existing:
                    existing.ota_room_code = ota_room_id
                    existing.ota_room_name = entry.otaRoomType or room_type.name
                    existing.is_active = True
                    existing.last_synced_at = datetime.utcnow()
                    await session.flush()
                    updated += 1
                else:
                    mapping = OTARoomMapping(
                        property_id=ota.property_id,
                        ota_connection_id=ota.id,
                        room_type_id=entry.pmsRoomTypeId,
                        ota_room_code=ota_room_id,
                        ota_room_name=entry.otaRoomType or room_type.name,
                        is_active=True,
                        sync_status="synced",
                        last_synced_at=datetime.utcnow()
                    )
                    session.add(mapping)
                    await session.flush()
                    created += 1
            except HTTPException:
                raise
            except Exception as e:
                errors.append({"pmsRoomTypeId": entry.pmsRoomTypeId, "error": str(e)})

        await session.commit()

        logger.info(f"[CHANNEL_MANAGER] POST /room-mappings/bulk - OTA: {ota_code}, created: {created}, updated: {updated}, errors: {len(errors)}")
        return {
            "success": True,
            "data": {
                "otaCode": ota_code,
                "created": created,
                "updated": updated,
                "errors": errors if errors else [],
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[CHANNEL_MANAGER] Bulk room mapping error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/room-mappings/{mapping_id}", tags=["Channel Manager"])
async def get_room_mapping(
    mapping_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get specific room mapping"""
    logger.info(f"[CHANNEL_MANAGER] GET /room-mappings/{mapping_id}")
    
    mapping = await session.get(OTARoomMapping, mapping_id)
    if not mapping:
        raise HTTPException(status_code=404, detail="Room mapping not found")
    
    room_type = await session.get(RoomType, mapping.room_type_id)
    ota_connection = await session.get(OTAConnection, mapping.ota_connection_id)
    
    # Get total room count for this room type
    from app.models.inventory import Room
    room_count_stmt = select(func.count(Room.id)).where(Room.room_type_id == mapping.room_type_id)
    room_count_result = await session.execute(room_count_stmt)
    total_rooms = room_count_result.scalar() or 0
    
    return {
        "success": True,
        "data": {
            "id": mapping.id,
            "pmsRoomType": room_type.name if room_type else "",
            "pmsRoomTypeId": mapping.room_type_id,
            "pmsRoomCode": room_type.slug if room_type else "",
            "basePrice": room_type.base_price if room_type else 0.0,
            "inventory": total_rooms,
            "otaMappings": [{
                "otaCode": ota_connection.ota_code if ota_connection else "",
                "otaRoomType": mapping.ota_room_name or mapping.ota_room_code,
                "otaRoomId": mapping.ota_room_code,
                "otaRoomCode": mapping.ota_room_code,
                "status": "active" if mapping.is_active else "inactive",
                "lastSync": mapping.last_synced_at.isoformat() if mapping.last_synced_at else None
            }]
        }
    }


@router.put("/room-mappings/{mapping_id}", tags=["Channel Manager"])
async def update_room_mapping(
    mapping_id: int,
    payload: Dict[str, Any] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Update room mapping"""
    logger.info(f"[CHANNEL_MANAGER] PUT /room-mappings/{mapping_id}")
    
    mapping = await session.get(OTARoomMapping, mapping_id)
    if not mapping:
        raise HTTPException(status_code=404, detail="Room mapping not found")
    
    # Update fields if provided
    if "otaRoomType" in payload:
        mapping.ota_room_name = payload["otaRoomType"]
    if "otaRoomId" in payload or "otaRoomCode" in payload:
        mapping.ota_room_code = payload.get("otaRoomId") or payload.get("otaRoomCode")
    if "status" in payload:
        mapping.is_active = payload["status"] == "active"
    
    mapping.last_synced_at = datetime.utcnow()
    await session.commit()
    await session.refresh(mapping)
    
    room_type = await session.get(RoomType, mapping.room_type_id)
    ota_connection = await session.get(OTAConnection, mapping.ota_connection_id)
    
    return {
        "success": True,
        "data": {
            "id": mapping.id,
            "pmsRoomType": room_type.name if room_type else "",
            "pmsRoomTypeId": mapping.room_type_id,
            "otaCode": ota_connection.ota_code if ota_connection else "",
            "otaRoomId": mapping.ota_room_code,
            "message": "Room mapping updated successfully"
        }
    }


@router.delete("/room-mappings/{mapping_id}", tags=["Channel Manager"])
async def delete_room_mapping(
    mapping_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Delete room mapping"""
    logger.info(f"[CHANNEL_MANAGER] DELETE /room-mappings/{mapping_id}")

    mapping = await session.get(OTARoomMapping, mapping_id)
    if not mapping:
        raise HTTPException(status_code=404, detail="Room mapping not found")

    await session.delete(mapping)
    await session.commit()

    return {"success": True, "message": "Room mapping deleted"}


@router.post("/room-mappings/validate", tags=["Channel Manager"])
async def validate_room_mapping(
    payload: Dict[str, Any] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Validate room mapping"""
    logger.info(f"[CHANNEL_MANAGER] POST /room-mappings/validate")
    
    pms_room_type_id = payload.get("pmsRoomTypeId")
    ota_code = payload.get("otaCode")
    
    if not pms_room_type_id or not ota_code:
        raise HTTPException(status_code=400, detail="Missing required fields: pmsRoomTypeId, otaCode")
    
    room_type = await session.get(RoomType, pms_room_type_id)
    if not room_type:
        return {
            "success": True,
            "data": {
                "valid": False,
                "errors": ["PMS room type not found"],
                "warnings": []
            }
        }
    
    # Get OTA connection
    ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
    ota_result = await session.execute(ota_stmt)
    ota = ota_result.scalar_one_or_none()
    
    if not ota:
        return {
            "success": True,
            "data": {
                "valid": False,
                "errors": [f"OTA with code {ota_code} not found"],
                "warnings": []
            }
        }
    
    # Check if mapping already exists
    existing_stmt = select(OTARoomMapping).where(
        and_(
            OTARoomMapping.ota_connection_id == ota.id,
            OTARoomMapping.room_type_id == pms_room_type_id
        )
    )
    existing_result = await session.execute(existing_stmt)
    existing = existing_result.scalar_one_or_none()
    
    warnings = []
    if existing and not existing.is_active:
        warnings.append("Mapping exists but is inactive")
    
    return {
        "success": True,
        "data": {
            "valid": True,
            "errors": [],
            "warnings": warnings
        }
    }


@router.post("/room-mappings/auto-map", tags=["Channel Manager"])
async def auto_map_rooms(
    request: Optional[AutoMapRequest] = Body(None),
    ota_code: Optional[str] = Query(None, alias="otaCode", description="OTA code (e.g. BOOKING, DUMMY). Use query param when request body is empty."),
    ota_connection_id: Optional[int] = Query(None, alias="otaConnectionId", description="OTA connection ID. Resolved to code when provided."),
    dry_run: bool = Query(False, alias="dryRun", description="If true, return suggestions only without applying mappings."),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Auto-map all PMS room types to OTA. Pass otaCode or otaConnectionId in query or body. Use dryRun=true for suggestions only."""
    try:
        # Resolve ota_code from body, query, or ota_connection_id lookup
        resolved_code = None
        if request:
            resolved_code = request.ota_code or request.otaCode
        if resolved_code:
            ota_code = resolved_code
        else:
            ota_code = (ota_code or "").strip() or None
        if not ota_code and ota_connection_id is not None:
            ota_row = await session.get(OTAConnection, ota_connection_id)
            if ota_row:
                ota_code = ota_row.ota_code
        if not ota_code:
            raise HTTPException(
                status_code=400,
                detail="Missing required parameter: provide otaCode (or ota_code) in query or body, or otaConnectionId"
            )
        
        user_info = current_user.email if current_user else "internal_service"
        logger.info(f"[CHANNEL_MANAGER] POST /room-mappings/auto-map - OTA: {ota_code}, User: {user_info}")
        print(f"[CHANNEL_MANAGER] Auto-mapping rooms for OTA: {ota_code}")
        
        # Get OTA connection
        ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
        ota_result = await session.execute(ota_stmt)
        ota = ota_result.scalar_one_or_none()
        
        if not ota:
            error_msg = f"OTA with code {ota_code} not found"
            logger.error(f"[CHANNEL_MANAGER] {error_msg}")
            raise HTTPException(status_code=404, detail=error_msg)
        
        # Get all room types
        room_types_stmt = select(RoomType)
        room_types_result = await session.execute(room_types_stmt)
        room_types = room_types_result.scalars().all()
        
        if not room_types:
            logger.warning(f"[CHANNEL_MANAGER] No room types found in system")
            return {
                "success": True,
                "data": {
                    "mappingsCreated": 0,
                    "suggestions": [],
                    "message": "No room types found to map"
                }
            }
        
        mappings_created = 0
        mappings_updated = 0
        suggestions = []
        errors = []
        
        for rt in room_types:
            try:
                # Check if mapping exists
                existing_stmt = select(OTARoomMapping).where(
                    and_(
                        OTARoomMapping.ota_connection_id == ota.id,
                        OTARoomMapping.room_type_id == rt.id
                    )
                )
                existing_result = await session.execute(existing_stmt)
                existing = existing_result.scalar_one_or_none()
                
                # Generate OTA room code (use slug if available, otherwise use name)
                room_slug = getattr(rt, 'slug', None) or rt.name.lower().replace(' ', '-')
                ota_room_code = f"{ota_code}-{room_slug.upper()}"
                
                if not dry_run:
                    if existing:
                        # Update existing mapping if needed
                        if existing.ota_room_code != ota_room_code or existing.ota_room_name != rt.name:
                            existing.ota_room_code = ota_room_code
                            existing.ota_room_name = rt.name
                            existing.is_active = True
                            existing.last_synced_at = datetime.utcnow()
                            mappings_updated += 1
                    else:
                        # Create new mapping
                        mapping = OTARoomMapping(
                            property_id=ota.property_id,
                            ota_connection_id=ota.id,
                            room_type_id=rt.id,
                            ota_room_code=ota_room_code,
                            ota_room_name=rt.name,
                            is_active=True,
                            sync_status="synced",
                            last_synced_at=datetime.utcnow()
                        )
                        session.add(mapping)
                        mappings_created += 1
                else:
                    # Dry run: count what would be created/updated
                    if existing:
                        if existing.ota_room_code != ota_room_code or existing.ota_room_name != rt.name:
                            mappings_updated += 1
                    else:
                        mappings_created += 1
                
                suggestions.append({
                    "pmsRoomType": rt.name,
                    "pmsRoomTypeId": rt.id,
                    "suggestedOTARoomType": rt.name,
                    "otaRoomCode": ota_room_code,
                    "confidence": 0.95
                })
            except Exception as e:
                error_msg = f"Error mapping room type {rt.name} (ID: {rt.id}): {str(e)}"
                logger.error(f"[CHANNEL_MANAGER] {error_msg}")
                errors.append(error_msg)
                continue
        
        if not dry_run:
            await session.commit()
        
        logger.info(f"[CHANNEL_MANAGER] Auto-mapping completed - Created: {mappings_created}, Updated: {mappings_updated}, Errors: {len(errors)}, dry_run={dry_run}")
        print(f"[CHANNEL_MANAGER] Auto-mapping completed - Created: {mappings_created}, Updated: {mappings_updated}")
        
        response_data = {
            "mappingsCreated": mappings_created,
            "mappingsUpdated": mappings_updated,
            "suggestions": suggestions,
            "totalRoomTypes": len(room_types),
            "successCount": mappings_created + mappings_updated,
            "errorCount": len(errors)
        }
        
        if errors:
            response_data["errors"] = errors
            response_data["message"] = f"Auto-mapping completed with {len(errors)} error(s)"
        elif dry_run:
            response_data["message"] = f"Dry run: {mappings_created + mappings_updated} mapping(s) would be created/updated. Call without dryRun to apply."
            response_data["dryRun"] = True
        else:
            response_data["message"] = f"Successfully auto-mapped {mappings_created + mappings_updated} room type(s)"
        
        return {
            "success": True,
            "data": response_data
        }
    
    except HTTPException:
        raise
    except Exception as e:
        error_msg = f"Unexpected error during auto-mapping: {str(e)}"
        logger.error(f"[CHANNEL_MANAGER] {error_msg}", exc_info=True)
        raise HTTPException(status_code=500, detail=error_msg)

# ============== RATE SYNC ENDPOINTS ==============

@router.get("/rates/calendar", tags=["Channel Manager"])
async def get_rate_calendar(
    start_date: date = Query(..., alias="startDate", description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., alias="endDate", description="End date (YYYY-MM-DD)"),
    room_type_id: Optional[int] = Query(None, alias="roomTypeId"),
    ota_code: Optional[str] = Query(None, alias="otaCode"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get rate calendar for date range"""
    logger.info(f"[CHANNEL_MANAGER] GET /rates/calendar - {start_date} to {end_date}, RoomType: {room_type_id}")
    
    # Get room types
    room_types_stmt = select(RoomType)
    if room_type_id:
        room_types_stmt = room_types_stmt.where(RoomType.id == room_type_id)
    room_types_result = await session.execute(room_types_stmt)
    room_types = room_types_result.scalars().all()
    
    calendar = {}
    current_date = start_date
    
    while current_date <= end_date:
        date_str = current_date.isoformat()
        calendar[date_str] = {}
        
        for rt in room_types:
            # Get base rate
            base_rate = rt.base_price
            
            # Get daily rate if exists
            daily_rate_stmt = select(DailyRate).where(
                and_(
                    DailyRate.room_type_id == rt.id,
                    DailyRate.date == current_date
                )
            ).limit(1)
            daily_rate_result = await session.execute(daily_rate_stmt)
            daily_rate = daily_rate_result.scalar_one_or_none()
            
            if daily_rate:
                base_rate = daily_rate.override_rate or daily_rate.base_rate
            
            # Get availability (use first() to handle potential duplicates)
            avail_stmt = select(DailyAvailability).where(
                and_(
                    DailyAvailability.room_type_id == rt.id,
                    DailyAvailability.date == current_date
                )
            ).limit(1)
            avail_result = await session.execute(avail_stmt)
            daily_avail = avail_result.scalar_one_or_none()
            
            # Get total room count
            from app.models.inventory import Room
            room_count_stmt = select(func.count(Room.id)).where(Room.room_type_id == rt.id)
            room_count_result = await session.execute(room_count_stmt)
            total_rooms = room_count_result.scalar() or 0
            
            available = total_rooms
            if daily_avail:
                available = daily_avail.available
            
            calendar[date_str][rt.name] = {
                "date": date_str,
                "roomType": rt.name,
                "rates": {"BAR": base_rate},
                "otaRates": {},
                "availability": available,
                "stopSell": daily_avail.is_closed if daily_avail else False,
                "cta": daily_avail.closed_to_arrival if daily_avail else False,
                "ctd": daily_avail.closed_to_departure if daily_avail else False
            }
        
        current_date += timedelta(days=1)
    
    logger.info(f"[CHANNEL_MANAGER] Generated rate calendar for {len(calendar)} dates")
    
    return {"success": True, "data": {"calendar": calendar}}


@router.put("/rates/calendar/{date}/{room_type}", tags=["Channel Manager"])
async def update_rate_calendar_entry(
    date: date = Path(..., description="Date (YYYY-MM-DD)"),
    room_type: str = Path(..., description="Room type name"),
    payload: Dict[str, Any] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Update rate for specific date and room type"""
    logger.info(f"[CHANNEL_MANAGER] PUT /rates/calendar/{date}/{room_type}")
    
    # Get room type
    room_type_stmt = select(RoomType).where(RoomType.name == room_type)
    room_type_result = await session.execute(room_type_stmt)
    rt = room_type_result.scalar_one_or_none()
    
    if not rt:
        raise HTTPException(status_code=404, detail=f"Room type {room_type} not found")
    
    # Update rates if provided
    if "rates" in payload:
        rates = payload["rates"]
        if "BAR" in rates:
            # Get or create daily rate
            daily_rate_stmt = select(DailyRate).where(
                and_(
                    DailyRate.room_type_id == rt.id,
                    DailyRate.date == date
                )
            ).limit(1)
            daily_rate_result = await session.execute(daily_rate_stmt)
            daily_rate = daily_rate_result.scalar_one_or_none()
            
            if not daily_rate:
                # Get default rate plan
                rate_plan_stmt = select(RatePlan).where(RatePlan.code == "BAR").limit(1)
                rate_plan_result = await session.execute(rate_plan_stmt)
                rate_plan = rate_plan_result.scalar_one_or_none()
                
                if rate_plan:
                    daily_rate = DailyRate(
                        room_type_id=rt.id,
                        rate_plan_id=rate_plan.id,
                        date=date,
                        base_rate=rt.base_price,
                        override_rate=rates["BAR"]
                    )
                    session.add(daily_rate)
                else:
                    logger.warning(f"[CHANNEL_MANAGER] BAR rate plan not found")
            else:
                daily_rate.override_rate = rates["BAR"]
                daily_rate.updated_at = datetime.utcnow()

        # Sync to DailyAvailability so CMS Availability Calendar and room views show the same price
        bar_rate = rates.get("BAR")
        if bar_rate is not None:
            avail_stmt = select(DailyAvailability).where(
                and_(
                    DailyAvailability.room_type_id == rt.id,
                    DailyAvailability.date == date
                )
            ).limit(1)
            avail_result = await session.execute(avail_stmt)
            daily_avail_for_rate = avail_result.scalar_one_or_none()
            if daily_avail_for_rate:
                daily_avail_for_rate.base_rate = bar_rate
                daily_avail_for_rate.updated_at = datetime.utcnow()
            else:
                room_count_stmt = select(func.count(Room.id)).where(Room.room_type_id == rt.id)
                room_count_result = await session.execute(room_count_stmt)
                total_rooms = room_count_result.scalar() or 0
                session.add(
                    DailyAvailability(
                        room_type_id=rt.id,
                        date=date,
                        total_rooms=total_rooms,
                        available=total_rooms,
                        base_rate=bar_rate,
                        is_closed=False,
                        closed_to_arrival=False,
                        closed_to_departure=False,
                    )
                )
    
    # Get or create daily availability (query FIRST, before checking payload keys)
    avail_stmt = select(DailyAvailability).where(
        and_(
            DailyAvailability.room_type_id == rt.id,
            DailyAvailability.date == date
        )
    ).limit(1)
    avail_result = await session.execute(avail_stmt)
    daily_avail = avail_result.scalar_one_or_none()

    # Update availability restrictions
    if "stopSell" in payload or "cta" in payload or "ctd" in payload:
        if not daily_avail:
            # Get total room count
            from app.models.inventory import Room
            room_count_stmt = select(func.count(Room.id)).where(Room.room_type_id == rt.id)
            room_count_result = await session.execute(room_count_stmt)
            total_rooms = room_count_result.scalar() or 0
            
            daily_avail = DailyAvailability(
                room_type_id=rt.id,
                date=date,
                total_rooms=total_rooms,
                available=total_rooms,
                is_closed=False,
                closed_to_arrival=False,
                closed_to_departure=False
            )
            session.add(daily_avail)
        
        if "stopSell" in payload:
            daily_avail.is_closed = payload["stopSell"]
        if "cta" in payload:
            daily_avail.closed_to_arrival = payload["cta"]
        if "ctd" in payload:
            daily_avail.closed_to_departure = payload["ctd"]
        
        daily_avail.updated_at = datetime.utcnow()
    
    # Update availability if provided
    if "availability" in payload:
        if not daily_avail:
            from app.models.inventory import Room
            room_count_stmt = select(func.count(Room.id)).where(Room.room_type_id == rt.id)
            room_count_result = await session.execute(room_count_stmt)
            total_rooms = room_count_result.scalar() or 0
            
            daily_avail = DailyAvailability(
                room_type_id=rt.id,
                date=date,
                total_rooms=total_rooms,
                available=payload["availability"],
                is_closed=False,
                closed_to_arrival=False,
                closed_to_departure=False
            )
            session.add(daily_avail)
        else:
            daily_avail.available = payload["availability"]
            daily_avail.updated_at = datetime.utcnow()
    
    await session.commit()
    
    logger.info(f"[CHANNEL_MANAGER] Updated rate calendar entry for {date}/{room_type}")
    
    # Return updated entry in same format as GET
    daily_rate_stmt = select(DailyRate).where(
        and_(
            DailyRate.room_type_id == rt.id,
            DailyRate.date == date
        )
    ).limit(1)
    daily_rate_result = await session.execute(daily_rate_stmt)
    daily_rate = daily_rate_result.scalar_one_or_none()
    
    avail_stmt = select(DailyAvailability).where(
        and_(
            DailyAvailability.room_type_id == rt.id,
            DailyAvailability.date == date
        )
    ).limit(1)
    avail_result = await session.execute(avail_stmt)
    daily_avail_final = avail_result.scalar_one_or_none()
    
    base_rate = rt.base_price
    if daily_rate:
        base_rate = daily_rate.override_rate or daily_rate.base_rate
    
    available = daily_avail_final.available if daily_avail_final else 0
    
    return {
        "success": True,
        "data": {
            "date": date.isoformat(),
            "roomType": rt.name,
            "rates": {"BAR": base_rate},
            "otaRates": {},
            "availability": available,
            "stopSell": daily_avail_final.is_closed if daily_avail_final else False,
            "cta": daily_avail_final.closed_to_arrival if daily_avail_final else False,
            "ctd": daily_avail_final.closed_to_departure if daily_avail_final else False
        }
    }


@router.post("/rates/push", tags=["Channel Manager"])
async def push_rates(
    payload: Dict[str, Any] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Push rates to OTAs"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] POST /rates/push - User: {user_info}")
    
    ota_codes = payload.get("otaCodes", [])
    date_range = payload.get("dateRange", {})
    room_type_ids = payload.get("roomTypeIds", [])
    
    logger.info(f"[CHANNEL_MANAGER] Pushing rates - OTAs: {ota_codes}, Dates: {date_range}")
    print(f"[CHANNEL_MANAGER] Pushing rates to OTAs: {ota_codes}")
    
    # Create real sync logs (ID + timestamp basis) for each OTA so Sync Logs tab shows actual entries
    created_logs: List[SyncLog] = []
    if ota_codes:
        stmt = select(OTAConnection).where(
            OTAConnection.ota_code.in_([c.upper() for c in ota_codes if c]),
            OTAConnection.is_active == True
        )
        result = await session.execute(stmt)
        otas = result.scalars().all()
        for ota in otas:
            sync_log = SyncLog(
                property_id=ota.property_id,
                ota_connection_id=ota.id,
                sync_type="rates",
                sync_direction="push",
                status="pending",
                started_at=datetime.utcnow()
            )
            session.add(sync_log)
            created_logs.append(sync_log)
        if created_logs:
            await session.commit()
            for log in created_logs:
                await session.refresh(log)
    
    first = created_logs[0] if created_logs else None
    return {
        "success": True,
        "data": {
            "syncId": first.id if first else None,
            "startedAt": first.started_at.isoformat() if first and first.started_at else None,
            "syncIds": [log.id for log in created_logs if log.id is not None],
            "status": "pending"
        }
    }


@router.post("/rates/pull", tags=["Channel Manager"])
async def pull_rates(
    payload: Dict[str, Any] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Pull rates from OTAs"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] POST /rates/pull - User: {user_info}")
    
    ota_codes = payload.get("otaCodes", [])
    date_range = payload.get("dateRange", {})
    room_type_ids = payload.get("roomTypeIds", [])
    
    logger.info(f"[CHANNEL_MANAGER] Pulling rates - OTAs: {ota_codes}, Dates: {date_range}")
    print(f"[CHANNEL_MANAGER] Pulling rates from OTAs: {ota_codes}")
    
    # Create real sync logs (ID + timestamp basis) for each OTA so Sync Logs tab shows actual entries
    created_logs: List[SyncLog] = []
    if ota_codes:
        stmt = select(OTAConnection).where(
            OTAConnection.ota_code.in_([c.upper() for c in ota_codes if c]),
            OTAConnection.is_active == True
        )
        result = await session.execute(stmt)
        otas = result.scalars().all()
        for ota in otas:
            sync_log = SyncLog(
                property_id=ota.property_id,
                ota_connection_id=ota.id,
                sync_type="rates",
                sync_direction="pull",
                status="pending",
                started_at=datetime.utcnow()
            )
            session.add(sync_log)
            created_logs.append(sync_log)
        if created_logs:
            await session.commit()
            for log in created_logs:
                await session.refresh(log)
    
    first = created_logs[0] if created_logs else None
    return {
        "success": True,
        "data": {
            "syncId": first.id if first else None,
            "startedAt": first.started_at.isoformat() if first and first.started_at else None,
            "syncIds": [log.id for log in created_logs if log.id is not None],
            "status": "pending"
        }
    }


@router.get("/rates/parity", tags=["Channel Manager"])
async def get_rate_parity(
    date: Optional[date] = Query(None, description="Date to check (default: today)"),
    threshold: Optional[float] = Query(10.0, description="Percentage threshold (default: 10)"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get rate parity issues"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /rates/parity - Date: {date}, Threshold: {threshold}%, User: {user_info}")
    
    if not date:
        date = date.today()
    
    # Get all room types
    room_types_stmt = select(RoomType)
    room_types_result = await session.execute(room_types_stmt)
    room_types = room_types_result.scalars().all()
    
    issues = []
    
    for rt in room_types:
        # Get rates for this room type on this date
        daily_rate_stmt = select(DailyRate).where(
            and_(
                DailyRate.room_type_id == rt.id,
                DailyRate.date == date
            )
        )
        daily_rate_result = await session.execute(daily_rate_stmt)
        daily_rates = daily_rate_result.scalars().all()
        
        if daily_rates:
            rates = [dr.override_rate or dr.base_rate for dr in daily_rates if dr.override_rate or dr.base_rate]
            if rates:
                min_rate = min(rates)
                max_rate = max(rates)
                if min_rate > 0:
                    difference = ((max_rate - min_rate) / min_rate) * 100
                    if difference > threshold:
                        issues.append({
                            "date": date.isoformat(),
                            "roomType": rt.name,
                            "minRate": min_rate,
                            "maxRate": max_rate,
                            "difference": round(difference, 2),
                            "otas": ["DUMMY"]  # Simplified - would check actual OTA rates
                        })
    
    logger.info(f"[CHANNEL_MANAGER] Found {len(issues)} rate parity issues")
    print(f"[CHANNEL_MANAGER] Rate parity check: {len(issues)} issues found")
    
    return {
        "success": True,
        "data": {
            "issues": issues
        }
    }

# ============== RESTRICTIONS ENDPOINTS ==============

@router.get("/restrictions", tags=["Channel Manager"])
async def get_restrictions(
    status: Optional[str] = Query(None, description="Filter by status: active, inactive, all"),
    room_type: Optional[str] = Query(None),
    ota_code: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get all restrictions"""
    logger.info(f"[CHANNEL_MANAGER] GET /restrictions - Status: {status}, RoomType: {room_type}")
    
    stmt = select(ChannelRestriction)
    
    if status == "active":
        stmt = stmt.where(ChannelRestriction.is_active == True)
    elif status == "inactive":
        stmt = stmt.where(ChannelRestriction.is_active == False)
    
    if room_type and room_type != "ALL":
        # Get room type ID
        rt_stmt = select(RoomType).where(RoomType.name == room_type)
        rt_result = await session.execute(rt_stmt)
        rt = rt_result.scalar_one_or_none()
        if rt:
            stmt = stmt.where(ChannelRestriction.room_type_id == rt.id)
    
    if date_from:
        stmt = stmt.where(ChannelRestriction.restriction_date >= date_from)
    if date_to:
        stmt = stmt.where(ChannelRestriction.restriction_date <= date_to)

    stmt = stmt.order_by(ChannelRestriction.id.desc())
    result = await session.execute(stmt)
    restrictions = result.scalars().all()

    # Group rows by group_id so each user-created restriction appears as one item
    # Rows without group_id (legacy) are treated individually
    groups = {}
    ungrouped = []
    for r in restrictions:
        if r.group_id:
            if r.group_id not in groups:
                groups[r.group_id] = []
            groups[r.group_id].append(r)
        else:
            ungrouped.append(r)

    items = []

    # Build grouped items — one entry per group_id
    for gid, rows in groups.items():
        first = rows[0]
        rt = await session.get(RoomType, first.room_type_id) if first.room_type_id else None
        ota = await session.get(OTAConnection, first.ota_connection_id) if first.ota_connection_id else None

        # Merge restriction types from all rows in the group
        min_stay = 1
        max_stay = None
        cta = False
        ctd = False
        stop_sell = False
        for r in rows:
            if r.restriction_type == "min_stay":
                min_stay = r.restriction_value
            elif r.restriction_type == "max_stay":
                max_stay = r.restriction_value
            elif r.restriction_type == "CTA":
                cta = True
            elif r.restriction_type == "CTD":
                ctd = True
            elif r.restriction_type == "stop_sell":
                stop_sell = True

        items.append({
            "id": first.id,
            "groupId": gid,
            "roomType": rt.name if rt else "ALL",
            "otaCode": ota.ota_code if ota else "ALL",
            "dateRange": {
                "start": (first.date_range_start or first.restriction_date).isoformat(),
                "end": (first.date_range_end or first.restriction_date).isoformat()
            },
            "restriction": {
                "minStay": min_stay,
                "maxStay": max_stay,
                "cta": cta,
                "ctd": ctd,
                "stopSell": stop_sell
            },
            "reason": first.reason,
            "isActive": first.is_active,
            "createdAt": first.created_at.isoformat() if first.created_at else None
        })

    # Legacy ungrouped rows (created before group_id was added)
    for r in ungrouped:
        rt = await session.get(RoomType, r.room_type_id) if r.room_type_id else None
        ota = await session.get(OTAConnection, r.ota_connection_id) if r.ota_connection_id else None
        items.append({
            "id": r.id,
            "groupId": None,
            "roomType": rt.name if rt else "ALL",
            "otaCode": ota.ota_code if ota else "ALL",
            "dateRange": {
                "start": (r.date_range_start or r.restriction_date).isoformat(),
                "end": (r.date_range_end or r.restriction_date).isoformat()
            },
            "restriction": {
                "minStay": r.restriction_value if r.restriction_type == "min_stay" else 1,
                "maxStay": r.restriction_value if r.restriction_type == "max_stay" else None,
                "cta": r.restriction_type == "CTA",
                "ctd": r.restriction_type == "CTD",
                "stopSell": r.restriction_type == "stop_sell"
            },
            "reason": r.reason,
            "isActive": r.is_active,
            "createdAt": r.created_at.isoformat() if r.created_at else None
        })

    # Sort by newest first (highest id)
    items.sort(key=lambda x: x["id"], reverse=True)

    logger.info(f"[CHANNEL_MANAGER] Found {len(items)} restriction groups")
    return {"success": True, "data": {"items": items, "total": len(items)}}


@router.post("/restrictions", tags=["Channel Manager"])
async def create_restriction(
    payload: RestrictionCreateRequest,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Create new restriction"""
    logger.info(f"[CHANNEL_MANAGER] POST /restrictions - RoomType: {payload.roomType}, OTA: {payload.otaCode}")

    # Get date range
    start_date = datetime.fromisoformat(payload.dateRange["start"]).date()
    end_date = datetime.fromisoformat(payload.dateRange["end"]).date()

    if start_date > end_date:
        raise HTTPException(
            status_code=400,
            detail="Start Date must be on or before End Date"
        )

    # Normalise roomType / otaCode into lists so we handle both single and multi-select
    room_type_names: List[str] = (
        ["ALL"] if payload.roomType == "ALL"
        else (payload.roomType if isinstance(payload.roomType, list) else [payload.roomType])
    )
    ota_codes: List[str] = (
        ["ALL"] if payload.otaCode == "ALL"
        else (payload.otaCode if isinstance(payload.otaCode, list) else [payload.otaCode])
    )

    # Resolve room type IDs
    room_type_ids: List[Optional[int]] = []
    for rt_name in room_type_names:
        if rt_name == "ALL":
            room_type_ids.append(None)
        else:
            rt_stmt = select(RoomType).where(RoomType.name == rt_name)
            rt_result = await session.execute(rt_stmt)
            rt = rt_result.scalar_one_or_none()
            if rt:
                room_type_ids.append(rt.id)
            else:
                raise HTTPException(status_code=404, detail=f"Room type {rt_name} not found")

    # Resolve OTA connection IDs
    ota_connection_ids: List[Optional[int]] = []
    for ota_code in ota_codes:
        if ota_code == "ALL":
            ota_connection_ids.append(None)
        else:
            ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
            ota_result = await session.execute(ota_stmt)
            ota = ota_result.scalar_one_or_none()
            if ota:
                ota_connection_ids.append(ota.id)
            else:
                raise HTTPException(status_code=404, detail=f"OTA {ota_code} not found")

    # Create restrictions for each room type × OTA × date × restriction type
    restriction = payload.restriction
    created_count = 0

    for room_type_id in room_type_ids:
        for ota_connection_id in ota_connection_ids:
            restriction_group_id = str(uuid.uuid4())

            current_date = start_date
            while current_date <= end_date:
                if restriction.get("minStay") and restriction["minStay"] > 1:
                    session.add(ChannelRestriction(
                        property_id=1, ota_connection_id=ota_connection_id, room_type_id=room_type_id,
                        restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                        group_id=restriction_group_id, restriction_type="min_stay",
                        restriction_value=restriction["minStay"], reason=payload.reason, is_active=True
                    ))
                    created_count += 1

                if restriction.get("maxStay"):
                    session.add(ChannelRestriction(
                        property_id=1, ota_connection_id=ota_connection_id, room_type_id=room_type_id,
                        restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                        group_id=restriction_group_id, restriction_type="max_stay",
                        restriction_value=restriction["maxStay"], reason=payload.reason, is_active=True
                    ))
                    created_count += 1

                if restriction.get("cta"):
                    session.add(ChannelRestriction(
                        property_id=1, ota_connection_id=ota_connection_id, room_type_id=room_type_id,
                        restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                        group_id=restriction_group_id, restriction_type="CTA",
                        restriction_value=1, reason=payload.reason, is_active=True
                    ))
                    created_count += 1

                if restriction.get("ctd"):
                    session.add(ChannelRestriction(
                        property_id=1, ota_connection_id=ota_connection_id, room_type_id=room_type_id,
                        restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                        group_id=restriction_group_id, restriction_type="CTD",
                        restriction_value=1, reason=payload.reason, is_active=True
                    ))
                    created_count += 1

                if restriction.get("stopSell"):
                    session.add(ChannelRestriction(
                        property_id=1, ota_connection_id=ota_connection_id, room_type_id=room_type_id,
                        restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                        group_id=restriction_group_id, restriction_type="stop_sell",
                        restriction_value=1, reason=payload.reason, is_active=True
                    ))
                    created_count += 1

                current_date += timedelta(days=1)

    if created_count == 0:
        raise HTTPException(
            status_code=400,
            detail="Please select at least one restriction type (CTA, CTD, Stop Sell, or set Min Stay > 1 / Max Stay)"
        )

    await session.commit()

    logger.info(f"[CHANNEL_MANAGER] Created {created_count} restriction entries")

    return {"success": True, "data": {"created": created_count, "message": "Restrictions created"}}


@router.get("/restrictions/{restriction_id}", tags=["Channel Manager"])
async def get_restriction(
    restriction_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get specific restriction"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /restrictions/{restriction_id} - User: {user_info}")
    
    restriction = await session.get(ChannelRestriction, restriction_id)
    if not restriction:
        raise HTTPException(status_code=404, detail="Restriction not found")
    
    rt = await session.get(RoomType, restriction.room_type_id) if restriction.room_type_id else None
    ota = await session.get(OTAConnection, restriction.ota_connection_id) if restriction.ota_connection_id else None
    
    return {
        "success": True,
        "data": {
            "id": restriction.id,
            "roomType": rt.name if rt else "ALL",
            "otaCode": ota.ota_code if ota else "ALL",
            "dateRange": {
                "start": restriction.restriction_date.isoformat(),
                "end": restriction.restriction_date.isoformat()
            },
            "restriction": {
                "minStay": restriction.restriction_value if restriction.restriction_type == "min_stay" else 1,
                "maxStay": restriction.restriction_value if restriction.restriction_type == "max_stay" else None,
                "cta": restriction.restriction_type == "CTA",
                "ctd": restriction.restriction_type == "CTD",
                "stopSell": restriction.restriction_type == "stop_sell"
            },
            "reason": restriction.reason,
            "isActive": restriction.is_active,
            "createdAt": restriction.created_at.isoformat() if restriction.created_at else None
        }
    }


@router.put("/restrictions/{restriction_id}", tags=["Channel Manager"])
async def update_restriction(
    restriction_id: int,
    payload: Dict[str, Any] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Update restriction — deletes old group rows and recreates with updated values"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] PUT /restrictions/{restriction_id} - User: {user_info}")

    try:
        restriction = await session.get(ChannelRestriction, restriction_id)
        if not restriction:
            raise HTTPException(status_code=404, detail="Restriction not found")

        # Find all rows in the same group
        if restriction.group_id:
            grp_stmt = select(ChannelRestriction).where(ChannelRestriction.group_id == restriction.group_id)
            grp_result = await session.execute(grp_stmt)
            group_rows = grp_result.scalars().all()
        else:
            group_rows = [restriction]

        # Capture old values before deleting
        old_start = restriction.date_range_start or restriction.restriction_date
        old_end = restriction.date_range_end or restriction.restriction_date
        old_reason = restriction.reason
        old_is_active = restriction.is_active
        old_room_type_id = restriction.room_type_id
        old_ota_connection_id = restriction.ota_connection_id
        old_group_id = restriction.group_id or str(uuid.uuid4())

        # Delete old group rows
        for row in group_rows:
            await session.delete(row)
        await session.flush()

        # Parse new values from payload, falling back to old values
        new_date_range = payload.get("dateRange")
        if new_date_range and "start" in new_date_range:
            raw_start = new_date_range["start"]
            start_date = date.fromisoformat(raw_start) if isinstance(raw_start, str) and len(raw_start) == 10 else datetime.fromisoformat(raw_start).date()
        else:
            start_date = old_start
        if new_date_range and "end" in new_date_range:
            raw_end = new_date_range["end"]
            end_date = date.fromisoformat(raw_end) if isinstance(raw_end, str) and len(raw_end) == 10 else datetime.fromisoformat(raw_end).date()
        else:
            end_date = old_end

        reason = payload.get("reason") if "reason" in payload else old_reason
        is_active = payload.get("isActive") if "isActive" in payload else old_is_active

        # Resolve room type (frontend may send a string or a list)
        room_type_id = old_room_type_id
        new_room_type = payload.get("roomType")
        if isinstance(new_room_type, list):
            new_room_type = new_room_type[0] if new_room_type else "ALL"
        if new_room_type == "ALL":
            room_type_id = None
        elif new_room_type:
            rt_stmt = select(RoomType).where(RoomType.name == new_room_type)
            rt_result = await session.execute(rt_stmt)
            rt = rt_result.scalar_one_or_none()
            if rt:
                room_type_id = rt.id

        # Resolve OTA connection (frontend may send a string or a list)
        ota_conn_id = old_ota_connection_id
        new_ota_code = payload.get("otaCode")
        if isinstance(new_ota_code, list):
            new_ota_code = new_ota_code[0] if new_ota_code else "ALL"
        if new_ota_code == "ALL":
            ota_conn_id = None
        elif new_ota_code:
            ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == new_ota_code)
            ota_result = await session.execute(ota_stmt)
            ota = ota_result.scalar_one_or_none()
            if ota:
                ota_conn_id = ota.id

        # Restriction types from payload
        new_restriction = payload.get("restriction", {})
        min_stay_val = new_restriction.get("minStay", 1)
        max_stay_val = new_restriction.get("maxStay")
        cta = new_restriction.get("cta", False)
        ctd = new_restriction.get("ctd", False)
        stop_sell = new_restriction.get("stopSell", False)

        # Recreate rows with updated values
        created_count = 0
        current_date = start_date
        while current_date <= end_date:
            if min_stay_val and min_stay_val > 0:
                session.add(ChannelRestriction(
                    property_id=1, ota_connection_id=ota_conn_id, room_type_id=room_type_id,
                    restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                    group_id=old_group_id, restriction_type="min_stay", restriction_value=min_stay_val,
                    reason=reason, is_active=is_active
                ))
                created_count += 1
            if max_stay_val and max_stay_val > 0:
                session.add(ChannelRestriction(
                    property_id=1, ota_connection_id=ota_conn_id, room_type_id=room_type_id,
                    restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                    group_id=old_group_id, restriction_type="max_stay", restriction_value=max_stay_val,
                    reason=reason, is_active=is_active
                ))
                created_count += 1
            if cta:
                session.add(ChannelRestriction(
                    property_id=1, ota_connection_id=ota_conn_id, room_type_id=room_type_id,
                    restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                    group_id=old_group_id, restriction_type="CTA", restriction_value=1,
                    reason=reason, is_active=is_active
                ))
                created_count += 1
            if ctd:
                session.add(ChannelRestriction(
                    property_id=1, ota_connection_id=ota_conn_id, room_type_id=room_type_id,
                    restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                    group_id=old_group_id, restriction_type="CTD", restriction_value=1,
                    reason=reason, is_active=is_active
                ))
                created_count += 1
            if stop_sell:
                session.add(ChannelRestriction(
                    property_id=1, ota_connection_id=ota_conn_id, room_type_id=room_type_id,
                    restriction_date=current_date, date_range_start=start_date, date_range_end=end_date,
                    group_id=old_group_id, restriction_type="stop_sell", restriction_value=1,
                    reason=reason, is_active=is_active
                ))
                created_count += 1
            current_date += timedelta(days=1)

        await session.commit()
        logger.info(f"[CHANNEL_MANAGER] Updated restriction group {old_group_id}, recreated {created_count} rows")

        return {"success": True, "data": {"groupId": old_group_id, "updated": created_count, "message": "Restriction updated"}}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[CHANNEL_MANAGER] Error updating restriction {restriction_id}: {e}", exc_info=True)
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update restriction: {str(e)}")


@router.put("/restrictions/{restriction_id}/toggle", tags=["Channel Manager"])
async def toggle_restriction(
    restriction_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Toggle restriction active status — toggles all rows in the group"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] PUT /restrictions/{restriction_id}/toggle - User: {user_info}")

    restriction = await session.get(ChannelRestriction, restriction_id)
    if not restriction:
        raise HTTPException(status_code=404, detail="Restriction not found")

    new_status = not restriction.is_active

    # Toggle all rows in the same group
    if restriction.group_id:
        grp_stmt = select(ChannelRestriction).where(ChannelRestriction.group_id == restriction.group_id)
        grp_result = await session.execute(grp_stmt)
        for row in grp_result.scalars().all():
            row.is_active = new_status
            row.updated_at = datetime.utcnow()
    else:
        restriction.is_active = new_status
        restriction.updated_at = datetime.utcnow()

    await session.commit()

    logger.info(f"[CHANNEL_MANAGER] Toggled restriction group {restriction.group_id or restriction_id} to {new_status}")

    return {"success": True, "data": {"isActive": new_status}}


@router.delete("/restrictions/{restriction_id}", tags=["Channel Manager"])
async def delete_restriction(
    restriction_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Delete restriction — deletes all rows in the group"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] DELETE /restrictions/{restriction_id} - User: {user_info}")

    restriction = await session.get(ChannelRestriction, restriction_id)
    if not restriction:
        raise HTTPException(status_code=404, detail="Restriction not found")

    # Delete all rows in the same group
    if restriction.group_id:
        grp_stmt = select(ChannelRestriction).where(ChannelRestriction.group_id == restriction.group_id)
        grp_result = await session.execute(grp_stmt)
        for row in grp_result.scalars().all():
            await session.delete(row)
    else:
        await session.delete(restriction)

    await session.commit()

    logger.info(f"[CHANNEL_MANAGER] Deleted restriction group: {restriction.group_id or restriction_id}")

    return {"success": True, "message": "Restriction deleted"}

# ============== SYNC LOGS ENDPOINTS ==============

@router.get("/sync-logs", tags=["Channel Manager"])
async def get_sync_logs(
    ota_code: Optional[str] = Query(None, description="Filter by OTA code (e.g. BOOKING, EXPEDIA)"),
    ota_connection_id: Optional[int] = Query(None, alias="otaConnectionId", description="Filter by OTA connection ID (use when navigating from View Logs for a specific OTA)"),
    action: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None, alias="dateFrom", description="Filter logs from this date (YYYY-MM-DD)"),
    date_to: Optional[date] = Query(None, alias="dateTo", description="Filter logs until this date (YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100, alias="pageSize"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get sync logs with pagination. Pass ota_connection_id when viewing logs for a specific OTA (e.g. from View Logs)."""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /sync-logs - Page: {page}, PageSize: {page_size}, ota_connection_id: {ota_connection_id}, ota_code: {ota_code}, dateFrom: {date_from}, dateTo: {date_to}, User: {user_info}")
    print(f"[CHANNEL_MANAGER] Getting sync logs - Page: {page}, PageSize: {page_size}, ota_connection_id: {ota_connection_id}")
    
    stmt = select(SyncLog)
    
    if ota_connection_id is not None:
        stmt = stmt.where(SyncLog.ota_connection_id == ota_connection_id)
    elif ota_code:
        ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
        ota_result = await session.execute(ota_stmt)
        ota = ota_result.scalar_one_or_none()
        if ota:
            stmt = stmt.where(SyncLog.ota_connection_id == ota.id)
    
    if date_from:
        stmt = stmt.where(SyncLog.started_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        stmt = stmt.where(SyncLog.started_at <= datetime.combine(date_to, datetime.max.time()))
    
    if action:
        # Map action types back to sync_type for filtering
        # Requirements action types: rate_update, availability_update, restriction_update, promotion_sync, booking_import, connection, bulk_sync
        action_to_sync_type = {
            "rate_update": "rates",
            "availability_update": "availability",
            "restriction_update": "restrictions",
            "promotion_sync": "promotions",
            "booking_import": "bookings",
            "connection": "connection",
            "bulk_sync": "full"
        }
        # Try exact match first, then partial match
        sync_type_filter = action_to_sync_type.get(action.lower())
        if sync_type_filter:
            stmt = stmt.where(func.lower(SyncLog.sync_type) == sync_type_filter.lower())
        else:
            # Fallback to partial match
            stmt = stmt.where(func.lower(SyncLog.sync_type).like(f"%{action.lower()}%"))
    
    if status:
        # Map requirements status to database status
        status_map = {
            "success": "success",
            "error": "failed",
            "warning": "partial",
            "pending": ["in_progress", "pending"]
        }
        mapped_status = status_map.get(status.lower())
        if mapped_status:
            if isinstance(mapped_status, list):
                stmt = stmt.where(SyncLog.status.in_(mapped_status))
            else:
                stmt = stmt.where(SyncLog.status == mapped_status)
        else:
            stmt = stmt.where(SyncLog.status == status)
    
    # Get total count
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total_result = await session.execute(count_stmt)
    total = total_result.scalar_one()
    
    # Apply pagination
    stmt = stmt.order_by(SyncLog.started_at.desc())
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    
    result = await session.execute(stmt)
    logs = result.scalars().all()
    
    # Map sync_type to action types according to requirements
    action_type_map = {
        "rates": "rate_update",
        "availability": "availability_update",
        "restrictions": "restriction_update",
        "promotions": "promotion_sync",
        "bookings": "booking_import",
        "full": "bulk_sync",
        "connection": "connection"
    }
    
    items = []
    for log in logs:
        ota = await session.get(OTAConnection, log.ota_connection_id)
        
        # Map sync_type to action type
        action_type = action_type_map.get(log.sync_type.lower(), log.sync_type.lower())
        
        # Map status
        status_map = {
            "success": "success",
            "failed": "error",
            "partial": "warning",
            "in_progress": "pending",
            "pending": "pending"
        }
        mapped_status = status_map.get(log.status.lower(), log.status.lower())
        
        # Build message
        if log.sync_type == "rates":
            message = f"Rates synced successfully for all room types" if mapped_status == "success" else f"Rate sync {mapped_status}"
        elif log.sync_type == "availability":
            message = f"Availability updated for {log.records_processed} room types" if mapped_status == "success" else f"Availability sync {mapped_status}"
        elif log.sync_type == "restrictions":
            message = f"Restrictions updated successfully" if mapped_status == "success" else f"Restriction sync {mapped_status}"
        elif log.sync_type == "bookings":
            message = f"Bookings imported successfully" if mapped_status == "success" else f"Booking import {mapped_status}"
        else:
            message = f"Sync {log.sync_type} - {mapped_status}"
        
        # Build details
        details = None
        if log.records_processed or log.records_failed or log.duration_seconds:
            details = {
                "recordsProcessed": log.records_processed,
                "recordsFailed": log.records_failed,
                "duration": log.duration_seconds
            }
            # Add additional details based on sync type
            if log.sync_type == "rates" and log.records_processed:
                details["changesCount"] = log.records_processed
            if log.sync_type == "availability" and log.records_processed:
                details["dateRange"] = "30 days"  # Would be calculated from actual dates
        
        items.append({
            "id": log.id,
            "timestamp": log.started_at.isoformat() if log.started_at else None,
            "date": log.started_at.date().isoformat() if log.started_at else None,
            "otaCode": ota.ota_code if ota else "",
            "otaName": ota.ota_name if ota else "",
            "action": action_type,
            "status": mapped_status,
            "message": message,
            "details": details
        })
    
    total_pages = (total + page_size - 1) // page_size
    
    logger.info(f"[CHANNEL_MANAGER] Found {len(items)} sync logs (Total: {total})")
    print(f"[CHANNEL_MANAGER] Found {len(items)} sync logs (Total: {total})")
    
    return {
        "success": True,
        "data": {
            "items": items,
            "total": total,
            "page": page,
            "pageSize": page_size,
            "totalPages": total_pages
        }
    }


@router.get("/sync-logs/export", tags=["Channel Manager"])
async def export_sync_logs(
    format: str = Query("csv", description="Export format: csv, excel, pdf"),
    ota_code: Optional[str] = Query(None),
    ota_connection_id: Optional[int] = Query(None, alias="otaConnectionId", description="Filter by OTA connection ID"),
    action: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None, alias="dateFrom"),
    date_to: Optional[date] = Query(None, alias="dateTo"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Export sync logs to CSV, Excel, or PDF"""
    logger.info(f"[CHANNEL_MANAGER] GET /sync-logs/export - Format: {format}, ota_connection_id: {ota_connection_id}, User: {current_user.email}")
    print(f"[CHANNEL_MANAGER] Exporting sync logs to {format.upper()} - User: {current_user.email}")
    
    # Create request hash to prevent duplicate exports
    request_params = f"{current_user.id}_{ota_code}_{ota_connection_id}_{action}_{status}_{date_from}_{date_to}_{format}"
    request_hash = hashlib.md5(request_params.encode()).hexdigest()
    cache_key = f"{current_user.id}_{request_hash}"
    
    # Check if this request was made recently
    current_time = time.time()
    if cache_key in _export_request_cache:
        last_request_time = _export_request_cache[cache_key]
        if current_time - last_request_time < _EXPORT_COOLDOWN_SECONDS:
            raise HTTPException(
                status_code=429,
                detail=f"Export request too frequent. Please wait {_EXPORT_COOLDOWN_SECONDS} seconds between exports."
            )
    
    # Update cache
    _export_request_cache[cache_key] = current_time
    
    # Clean old cache entries (older than cooldown period)
    _export_request_cache.clear()  # Simple cleanup - in production, use TTL-based cache
    
    # Use same filtering logic as get_sync_logs
    stmt = select(SyncLog)
    
    if ota_connection_id is not None:
        stmt = stmt.where(SyncLog.ota_connection_id == ota_connection_id)
    elif ota_code:
        ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
        ota_result = await session.execute(ota_stmt)
        ota = ota_result.scalar_one_or_none()
        if ota:
            stmt = stmt.where(SyncLog.ota_connection_id == ota.id)
    
    if action:
        action_to_sync_type = {
            "rate_update": "rates",
            "availability_update": "availability",
            "restriction_update": "restrictions",
            "promotion_sync": "promotions",
            "booking_import": "bookings",
            "connection": "connection",
            "bulk_sync": "full"
        }
        sync_type_filter = action_to_sync_type.get(action.lower())
        if sync_type_filter:
            stmt = stmt.where(func.lower(SyncLog.sync_type) == sync_type_filter.lower())
        else:
            stmt = stmt.where(func.lower(SyncLog.sync_type).like(f"%{action.lower()}%"))
    
    if status:
        status_map = {
            "success": "success",
            "error": "failed",
            "warning": "partial",
            "pending": ["in_progress", "pending"]
        }
        mapped_status = status_map.get(status.lower())
        if mapped_status:
            if isinstance(mapped_status, list):
                stmt = stmt.where(SyncLog.status.in_(mapped_status))
            else:
                stmt = stmt.where(SyncLog.status == mapped_status)
        else:
            stmt = stmt.where(SyncLog.status == status)
    
    if date_from:
        stmt = stmt.where(SyncLog.started_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        stmt = stmt.where(SyncLog.started_at <= datetime.combine(date_to, datetime.max.time()))
    
    stmt = stmt.order_by(SyncLog.started_at.desc())
    result = await session.execute(stmt)
    logs = result.scalars().all()

    # Prepare data
    export_data = []
    for log in logs:
        ota = await session.get(OTAConnection, log.ota_connection_id)
        export_data.append({
            "timestamp": log.started_at.isoformat() if log.started_at else "",
            "ota_code": ota.ota_code if ota else "",
            "ota_name": ota.ota_name if ota else "",
            "action": log.sync_type,
            "status": log.status,
            "records_processed": log.records_processed or 0,
            "records_failed": log.records_failed or 0,
            "duration": log.duration_seconds or 0
        })
    
    # Generate export based on format
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    
    if format.lower() == "csv":
        return _export_csv(export_data, timestamp)
    elif format.lower() in ["excel", "xlsx"]:
        return _export_excel(export_data, timestamp)
    elif format.lower() == "pdf":
        return _export_pdf(export_data, timestamp)
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported format: {format}. Supported formats: csv, excel, pdf"
        )


@router.get("/sync-logs/{log_id}", tags=["Channel Manager"])
async def get_sync_log(
    log_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get specific sync log with details"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /sync-logs/{log_id} - User: {user_info}")
    print(f"[CHANNEL_MANAGER] Getting sync log: {log_id}")
    
    log = await session.get(SyncLog, log_id)
    if not log:
        raise HTTPException(status_code=404, detail="Sync log not found")
    
    ota = await session.get(OTAConnection, log.ota_connection_id)
    
    err_details = log.error_details
    err_msg = None
    if err_details is not None:
        if isinstance(err_details, dict) and "message" in err_details:
            err_msg = str(err_details["message"])
        elif isinstance(err_details, str):
            err_msg = err_details
        else:
            err_msg = str(err_details) if err_details else None

    # Map sync_type to action type
    action_type_map = {
        "rates": "rate_update",
        "availability": "availability_update",
        "restrictions": "restriction_update",
        "promotions": "promotion_sync",
        "bookings": "booking_import",
        "full": "bulk_sync",
        "connection": "connection"
    }
    action_type = action_type_map.get(log.sync_type.lower(), log.sync_type.lower())
    
    # Map status
    status_map = {
        "success": "success",
        "failed": "error",
        "partial": "warning",
        "in_progress": "pending",
        "pending": "pending"
    }
    mapped_status = status_map.get(log.status.lower(), log.status.lower())
    
    return {
        "success": True,
        "data": {
            "id": log.id,
            "timestamp": log.started_at.isoformat() if log.started_at else None,
            "otaCode": ota.ota_code if ota else "",
            "otaName": ota.ota_name if ota else "",
            "action": action_type,
            "status": mapped_status,
            "message": f"Sync {log.sync_type} - {mapped_status}",
            "details": {
                "recordsProcessed": log.records_processed,
                "recordsFailed": log.records_failed,
                "duration": log.duration_seconds,
                "startedAt": log.started_at.isoformat() if log.started_at else None,
                "completedAt": log.completed_at.isoformat() if log.completed_at else None,
                "errorMessage": err_msg,
                "errorDetails": err_details
            }
        }
    }


@router.delete("/sync-logs", tags=["Channel Manager"])
async def clear_sync_logs(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Clear all sync logs (admin only)"""
    logger.info(f"[CHANNEL_MANAGER] DELETE /sync-logs - User: {current_user.email}")
    print(f"[CHANNEL_MANAGER] Clearing all sync logs - User: {current_user.email}")
    
    # Delete all sync logs
    stmt = select(SyncLog)
    result = await session.execute(stmt)
    logs = result.scalars().all()
    
    for log in logs:
        await session.delete(log)
    
    await session.commit()
    
    logger.info(f"[CHANNEL_MANAGER] Cleared {len(logs)} sync logs")
    print(f"[CHANNEL_MANAGER] Cleared {len(logs)} sync logs")
    
    return {"success": True, "message": "Sync logs cleared"}


def _export_csv(data: List[Dict[str, Any]], timestamp: str) -> Response:
    """Export data to CSV format"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(["Timestamp", "OTA Code", "OTA Name", "Action", "Status", "Records Processed", "Records Failed", "Duration (s)"])
    
    # Write data
    for row in data:
        writer.writerow([
            row["timestamp"],
            row["ota_code"],
            row["ota_name"],
            row["action"],
            row["status"],
            row["records_processed"],
            row["records_failed"],
            row["duration"]
        ])
    
    csv_content = output.getvalue()
    output.close()
    
    logger.info(f"[CHANNEL_MANAGER] Exported {len(data)} sync logs to CSV")
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=sync-logs-{timestamp}.csv"}
    )


def _export_excel(data: List[Dict[str, Any]], timestamp: str) -> StreamingResponse:
    """Export data to Excel format"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Please install openpyxl: pip install openpyxl"
        )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sync Logs"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Sync Logs Export"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    
    # Generated timestamp
    ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws.merge_cells('A2:H2')
    
    # Blank row
    current_row = 4
    
    # Headers
    headers = ["Timestamp", "OTA Code", "OTA Name", "Action", "Status", "Records Processed", "Records Failed", "Duration (s)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    current_row += 1
    
    # Data rows
    for row_data in data:
        ws.cell(row=current_row, column=1, value=row_data["timestamp"]).border = thin_border
        ws.cell(row=current_row, column=2, value=row_data["ota_code"]).border = thin_border
        ws.cell(row=current_row, column=3, value=row_data["ota_name"]).border = thin_border
        ws.cell(row=current_row, column=4, value=row_data["action"]).border = thin_border
        ws.cell(row=current_row, column=5, value=row_data["status"]).border = thin_border
        ws.cell(row=current_row, column=6, value=row_data["records_processed"]).border = thin_border
        ws.cell(row=current_row, column=7, value=row_data["records_failed"]).border = thin_border
        ws.cell(row=current_row, column=8, value=row_data["duration"]).border = thin_border
        current_row += 1
    
    # Adjust column widths
    column_widths = [25, 12, 20, 15, 12, 18, 16, 15]
    for col, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    logger.info(f"[CHANNEL_MANAGER] Exported {len(data)} sync logs to Excel")
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sync-logs-{timestamp}.xlsx"}
    )


def _export_pdf(data: List[Dict[str, Any]], timestamp: str) -> StreamingResponse:
    """Export data to PDF format"""
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="PDF export not available. Please install reportlab: pip install reportlab"
        )
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    elements.append(Paragraph("<b>Sync Logs Export</b>", styles['Title']))
    elements.append(Spacer(1, 12))
    
    # Generated timestamp
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Summary
    elements.append(Paragraph(f"Total Records: {len(data)}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Table data
    table_data = [["Timestamp", "OTA Code", "OTA Name", "Action", "Status", "Records", "Failed", "Duration"]]
    
    for row_data in data:
        # Truncate timestamp for display
        timestamp_display = row_data["timestamp"][:19] if len(row_data["timestamp"]) > 19 else row_data["timestamp"]
        table_data.append([
            timestamp_display,
            row_data["ota_code"],
            row_data["ota_name"][:20] if len(row_data["ota_name"]) > 20 else row_data["ota_name"],
            row_data["action"],
            row_data["status"],
            str(row_data["records_processed"]),
            str(row_data["records_failed"]),
            f"{row_data['duration']:.2f}" if row_data["duration"] else "0.00"
        ])
    
    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    logger.info(f"[CHANNEL_MANAGER] Exported {len(data)} sync logs to PDF")
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=sync-logs-{timestamp}.pdf"}
    )

# ============== CMS AVAILABILITY ENDPOINT ==============

@router.put("/cms/availability/bulk-update", tags=["Channel Manager"])
async def cms_bulk_update_availability(
    updates: List[Dict[str, Any]] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """
    Bulk update availability for channel manager.
    This endpoint is called by the dummy channel manager to update Glimmora backend.
    Allows internal service calls without authentication.
    """
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] PUT /cms/availability/bulk-update - {len(updates)} updates, User: {user_info}")
    
    # Import here to avoid circular imports
    from app.api.v1.availability import BulkAvailabilityUpdate, ensure_daily_availability_record
    
    updated_count = 0
    
    for update_data in updates:
        try:
            # Convert to BulkAvailabilityUpdate format
            update = BulkAvailabilityUpdate(
                room_type_id=update_data.get("room_type_id"),
                start_date=datetime.fromisoformat(update_data.get("start_date")).date() if isinstance(update_data.get("start_date"), str) else update_data.get("start_date"),
                end_date=datetime.fromisoformat(update_data.get("end_date")).date() if isinstance(update_data.get("end_date"), str) else update_data.get("end_date"),
                is_closed=update_data.get("is_closed"),
                min_stay=update_data.get("min_stay"),
                max_stay=update_data.get("max_stay"),
                closed_to_arrival=update_data.get("closed_to_arrival"),
                closed_to_departure=update_data.get("closed_to_departure")
            )
            
            current_date = update.start_date
            while current_date <= update.end_date:
                daily_avail = await ensure_daily_availability_record(
                    session, update.room_type_id, current_date
                )
                
                if update.is_closed is not None:
                    daily_avail.is_closed = update.is_closed
                if update.min_stay is not None:
                    daily_avail.min_stay = update.min_stay
                if update.max_stay is not None:
                    daily_avail.max_stay = update.max_stay
                if update.closed_to_arrival is not None:
                    daily_avail.closed_to_arrival = update.closed_to_arrival
                if update.closed_to_departure is not None:
                    daily_avail.closed_to_departure = update.closed_to_departure
                
                daily_avail.updated_at = datetime.utcnow()
                updated_count += 1
                current_date += timedelta(days=1)
        except Exception as e:
            logger.error(f"[CHANNEL_MANAGER] Error updating availability: {e}")
            continue
    
    await session.commit()
    
    logger.info(f"[CHANNEL_MANAGER] Updated {updated_count} availability records")
    print(f"[CHANNEL_MANAGER] Bulk availability update completed: {updated_count} records updated")
    
    # Broadcast SSE event for real-time frontend updates
    try:
        from app.api.v1.webhooks import broadcast_sse_event
        asyncio.create_task(
            broadcast_sse_event(
                "availability.updated",
                {
                    "updated_count": updated_count,
                    "ota_connection_id": None  # Bulk update from channel manager
                }
            )
        )
        logger.info(f"[CHANNEL_MANAGER] Broadcasted SSE event for availability update")
    except Exception as sse_error:
        logger.warning(f"[CHANNEL_MANAGER] Failed to broadcast SSE event: {sse_error}")
    
    return {
        "success": True,
        "updated_records": updated_count,
        "message": f"Successfully updated {updated_count} availability records"
    }


# ============== STATS ENDPOINTS ==============

@router.get("/stats", tags=["Channel Manager"])
async def get_channel_stats(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get channel manager statistics"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /stats - User: {user_info}")
    print(f"[CHANNEL_MANAGER] Getting channel stats")
    
    # Get OTA connections
    ota_stmt = select(OTAConnection)
    ota_result = await session.execute(ota_stmt)
    otas = ota_result.scalars().all()
    
    connected_count = sum(1 for ota in otas if ota.connection_status == "connected")
    disconnected_count = sum(1 for ota in otas if ota.connection_status == "disconnected")
    error_count = sum(1 for ota in otas if ota.connection_status == "error")
    
    # Get list of OTAs with errors for direct visibility
    error_otas_list = [
        {
            "id": ota.id,
            "name": ota.ota_name,
            "code": ota.ota_code,
            "errorMessage": ota.error_message,
            "logo": ota.logo_url,
            "color": ota.brand_color
        }
        for ota in otas if ota.connection_status == "error"
    ]
    
    # Get bookings from OTAs
    bookings_stmt = select(Booking).where(Booking.booking_source.isnot(None))
    bookings_result = await session.execute(bookings_stmt)
    bookings = bookings_result.scalars().all()
    
    total_bookings = len([b for b in bookings if b.status != "cancelled"])
    total_revenue = sum(b.total_price for b in bookings if b.status != "cancelled")
    
    # Get room types
    room_types_stmt = select(RoomType)
    room_types_result = await session.execute(room_types_stmt)
    room_types = room_types_result.scalars().all()
    
    # Get mapped room types
    mappings_stmt = select(OTARoomMapping).where(OTARoomMapping.is_active == True)
    mappings_result = await session.execute(mappings_stmt)
    mappings = mappings_result.scalars().all()
    mapped_room_types = len(set(m.room_type_id for m in mappings))
    
    # Get active restrictions
    restrictions_stmt = select(ChannelRestriction).where(ChannelRestriction.is_active == True)
    restrictions_result = await session.execute(restrictions_stmt)
    active_restrictions = len(restrictions_result.scalars().all())
    
    # Fetch ALL reviews once (avoids N+1 per-OTA review queries)
    all_reviews_result = await session.execute(select(Review))
    all_reviews = all_reviews_result.scalars().all()

    source_map = {
        "BOOKING": "booking_com",
        "EXPEDIA": "expedia",
        "AGODA": "agoda",
        "AIRBNB": "airbnb",
        "TRIPADVISOR": "tripadvisor"
    }

    # Calculate channel performance
    channel_performance = []
    for ota in otas:
        # Filter bookings for this OTA using helper function
        ota_bookings = filter_bookings_by_ota(bookings, ota.ota_code)
        ota_revenue = sum(b.total_price for b in ota_bookings if b.status != "cancelled")

        # Match reviews in-memory
        review_source = source_map.get(ota.ota_code.upper(), ota.ota_code.lower())
        ota_code_lower = ota.ota_code.lower()
        ota_name_lower = ota.ota_name.lower()
        ota_reviews = [
            r for r in all_reviews
            if r.source == review_source
            or (r.source and ota_code_lower in r.source.lower())
            or (r.source and ota_name_lower in r.source.lower())
        ]
        avg_rating = sum(r.overall_rating for r in ota_reviews) / len(ota_reviews) if ota_reviews else 0.0

        channel_performance.append({
            "id": ota.id,
            "name": ota.ota_name,
            "code": ota.ota_code,
            "color": ota.brand_color or "#000000",
            "bookings": {
                "count": len(ota_bookings),
                "drillDownUrl": f"/api/v1/channel-manager/performance/{ota.ota_code}/bookings",
                "clickable": True
            },
            "revenue": {
                "amount": ota_revenue,
                "drillDownUrl": f"/api/v1/channel-manager/performance/{ota.ota_code}/revenue",
                "clickable": True
            },
            "rating": {
                "value": round(avg_rating, 2),
                "totalReviews": len(ota_reviews),
                "drillDownUrl": f"/api/v1/channel-manager/performance/{ota.ota_code}/reviews",
                "clickable": True
            },
            "commission": ota.commission_rate or 0.0,
            "conversionRate": 0.0
        })
    
    avg_commission = sum(ota.commission_rate or 0.0 for ota in otas) / len(otas) if otas else 0.0
    
    # Calculate revenue and bookings trends (last 7 days)
    today = date.today()
    revenue_trend = []
    bookings_trend = []
    for i in range(6, -1, -1):  # Last 7 days
        trend_date = today - timedelta(days=i)
        day_bookings = [
            b for b in bookings
            if b.created_at and b.created_at.date() == trend_date
        ]
        day_revenue = sum(b.total_price for b in day_bookings if b.status != "cancelled")
        revenue_trend.append(day_revenue)
        bookings_trend.append(len([b for b in day_bookings if b.status != "cancelled"]))
    
    # Calculate average rate (ADR)
    confirmed_bookings = [b for b in bookings if b.status != "cancelled"]
    avg_rate = sum(b.total_price / b.nights for b in confirmed_bookings if b.nights and b.nights > 0) / len(confirmed_bookings) if confirmed_bookings else 0.0
    
    # Calculate occupancy rate
    # Get total available room nights in last 30 days
    from app.models.inventory import Room
    rooms_stmt = select(Room)
    rooms_result = await session.execute(rooms_stmt)
    all_rooms = rooms_result.scalars().all()
    total_room_nights = len(all_rooms) * 30  # Approximate for last 30 days
    
    # Calculate sold room nights
    sold_room_nights = sum(b.nights for b in confirmed_bookings if b.nights)
    occupancy_rate = (sold_room_nights / total_room_nights * 100) if total_room_nights > 0 else 0.0
    
    # Get last sync time (most recent sync across all OTAs)
    last_sync = None
    for ota in otas:
        if ota.last_sync_at:
            if last_sync is None or ota.last_sync_at > last_sync:
                last_sync = ota.last_sync_at
    
    # Calculate growth percentages (compare last 7 days to previous 7 days)
    prev_7_days_revenue = sum(revenue_trend[:3]) if len(revenue_trend) >= 3 else 0.0
    curr_7_days_revenue = sum(revenue_trend[-3:]) if len(revenue_trend) >= 3 else 0.0
    revenue_growth = "+0%"
    if prev_7_days_revenue > 0:
        growth_pct = ((curr_7_days_revenue - prev_7_days_revenue) / prev_7_days_revenue) * 100
        revenue_growth = f"{'+' if growth_pct >= 0 else ''}{growth_pct:.1f}%"
    
    prev_7_days_bookings = sum(bookings_trend[:3]) if len(bookings_trend) >= 3 else 0
    curr_7_days_bookings = sum(bookings_trend[-3:]) if len(bookings_trend) >= 3 else 0
    bookings_growth = "+0%"
    if prev_7_days_bookings > 0:
        growth_pct = ((curr_7_days_bookings - prev_7_days_bookings) / prev_7_days_bookings) * 100
        bookings_growth = f"{'+' if growth_pct >= 0 else ''}{growth_pct:.1f}%"
    
    # Get rate parity issues (single query for all room types today)
    rate_parity_issues = []
    today_date = date.today()
    check_rt_ids = [rt.id for rt in room_types[:5]]
    if check_rt_ids:
        all_daily_rates_result = await session.execute(
            select(DailyRate).where(
                and_(
                    DailyRate.room_type_id.in_(check_rt_ids),
                    DailyRate.date == today_date
                )
            )
        )
        all_daily_rates = all_daily_rates_result.scalars().all()
        # Group by room_type_id
        rates_by_rt = {}
        for dr in all_daily_rates:
            rates_by_rt.setdefault(dr.room_type_id, []).append(dr)

        rt_lookup = {rt.id: rt for rt in room_types[:5]}
        for rt_id, daily_rates in rates_by_rt.items():
            rates = [dr.override_rate or dr.base_rate for dr in daily_rates if dr.override_rate or dr.base_rate]
            if rates and len(rates) > 1:
                min_rate = min(rates)
                max_rate = max(rates)
                if min_rate > 0:
                    difference = max_rate - min_rate
                    if difference > 0:
                        rt = rt_lookup.get(rt_id)
                        rate_parity_issues.append({
                            "roomType": rt.name if rt else f"Room Type {rt_id}",
                            "minRate": min_rate,
                            "maxRate": max_rate,
                            "difference": round(difference, 2)
                        })
    
    logger.info(f"[CHANNEL_MANAGER] Calculated stats - Connected: {connected_count}, Bookings: {total_bookings}, Revenue: {total_revenue}")
    print(f"[CHANNEL_MANAGER] Channel stats - Connected: {connected_count}, Bookings: {total_bookings}, Revenue: ${total_revenue:.2f}")
    
    return {
        "success": True,
        "data": {
            "connectedOTAs": connected_count,
            "disconnectedOTAs": disconnected_count,
            "errorOTAs": error_count,
            "errorOTAsList": error_otas_list,  # List of OTAs with errors for direct visibility
            "totalBookings": total_bookings,
            "totalRevenue": total_revenue,
            "mappedRoomTypes": mapped_room_types,
            "totalRoomTypes": len(room_types),
            "activeRestrictions": active_restrictions,
            "rateParityIssues": rate_parity_issues,
            "lastSync": last_sync.isoformat() if last_sync else None,
            "revenueTrend": revenue_trend,
            "bookingsTrend": bookings_trend,
            "channelPerformance": channel_performance,
            "avgCommission": avg_commission,
            "avgConversionRate": 0.0,
            "revenueGrowth": revenue_growth,
            "bookingsGrowth": bookings_growth,
            "avgRate": round(avg_rate, 2),
            "occupancyRate": round(occupancy_rate, 1)
        }
    }


@router.get("/performance", tags=["Channel Manager"])
@router.get("/performance/detailed", tags=["Channel Manager"])
async def get_ota_performance_detailed(
    start_date: Optional[date] = Query(None, alias="startDate", description="Start date for analysis (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, alias="endDate", description="End date for analysis (YYYY-MM-DD)"),
    ota_code: Optional[str] = Query(None, alias="otaCode", description="Filter by specific OTA code"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get detailed OTA performance data - used for 'View All' page"""
    try:
        user_info = current_user.email if current_user else "internal_service"
        logger.info(f"[CHANNEL_MANAGER] GET /performance/detailed - User: {user_info}, OTA: {ota_code}, DateRange: {start_date} to {end_date}")
        print(f"[CHANNEL_MANAGER] Getting detailed OTA performance data")
        
        # Set default date range (last 30 days if not specified)
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # Get OTA connections
        ota_stmt = select(OTAConnection)
        if ota_code:
            ota_stmt = ota_stmt.where(OTAConnection.ota_code == ota_code)
        ota_result = await session.execute(ota_stmt)
        otas = ota_result.scalars().all()
        
        if not otas:
            return {
                "success": True,
                "data": {
                    "period": {
                        "start": start_date.isoformat(),
                        "end": end_date.isoformat()
                    },
                    "channels": [],
                    "summary": {
                        "totalChannels": 0,
                        "totalBookings": 0,
                        "totalRevenue": 0.0,
                        "totalCommission": 0.0,
                        "netRevenue": 0.0
                    }
                }
            }
        
        # Get bookings in date range
        bookings_stmt = select(Booking).where(
            and_(
                Booking.booking_source.isnot(None),
                Booking.created_at >= datetime.combine(start_date, datetime.min.time()),
                Booking.created_at <= datetime.combine(end_date, datetime.max.time())
            )
        )
        bookings_result = await session.execute(bookings_stmt)
        all_bookings = bookings_result.scalars().all()
        
        # Calculate performance for each OTA
        channel_performance = []
        total_bookings = 0
        total_revenue = 0.0
        total_commission = 0.0
        
        for ota in otas:
            # Filter bookings for this OTA using helper function
            ota_bookings = filter_bookings_by_ota(all_bookings, ota.ota_code)
            
            # Separate by status
            confirmed_bookings = [b for b in ota_bookings if b.status in ["confirmed", "checked_in", "checked_out"]]
            cancelled_bookings = [b for b in ota_bookings if b.status == "cancelled"]
            pending_bookings = [b for b in ota_bookings if b.status == "pending"]
            
            # Calculate revenue
            confirmed_revenue = sum(b.total_price for b in confirmed_bookings)
            cancelled_revenue = sum(b.total_price for b in cancelled_bookings)
            net_revenue = confirmed_revenue - cancelled_revenue
            
            # Calculate commission
            commission_rate = ota.commission_rate or 0.0
            commission_amount = (confirmed_revenue * commission_rate) / 100
            net_revenue_after_commission = net_revenue - commission_amount
            
            # Get reviews for rating calculation
            source_map = {
                "BOOKING": "booking_com",
                "EXPEDIA": "expedia",
                "AGODA": "agoda",
                "AIRBNB": "airbnb",
                "TRIPADVISOR": "tripadvisor"
            }
            review_source = source_map.get(ota.ota_code.upper(), ota.ota_code.lower())
            reviews_stmt = select(Review).where(
                and_(
                    or_(
                        Review.source == review_source,
                        Review.source.like(f"%{ota.ota_code.lower()}%"),
                        Review.source.like(f"%{ota.ota_name.lower()}%")
                    ),
                    Review.review_date >= datetime.combine(start_date, datetime.min.time()),
                    Review.review_date <= datetime.combine(end_date, datetime.max.time())
                )
            )
            reviews_result = await session.execute(reviews_stmt)
            ota_reviews = reviews_result.scalars().all()
            avg_rating = sum(r.overall_rating for r in ota_reviews) / len(ota_reviews) if ota_reviews else 0.0
            
            # Calculate trends (last 7 days)
            today = date.today()
            revenue_trend = []
            bookings_trend = []
            
            for i in range(6, -1, -1):  # Last 7 days
                trend_date = today - timedelta(days=i)
                day_bookings = [
                    b for b in ota_bookings
                    if b.created_at and b.created_at.date() == trend_date
                ]
                day_revenue = sum(b.total_price for b in day_bookings if b.status != "cancelled")
                revenue_trend.append(day_revenue)
                bookings_trend.append(len([b for b in day_bookings if b.status != "cancelled"]))
            
            # Calculate conversion rate (if we have data)
            conversion_rate = 0.0  # Would need impression data to calculate properly
            
            # Calculate average booking value
            avg_booking_value = confirmed_revenue / len(confirmed_bookings) if confirmed_bookings else 0.0
            
            # Get last sync time
            last_sync = ota.last_sync_at.isoformat() if ota.last_sync_at else None
            
            channel_data = {
                "id": ota.id,
                "name": ota.ota_name,
                "code": ota.ota_code,
                "logo": ota.logo_url,
                "color": ota.brand_color or "#000000",
                "status": ota.connection_status,
                "lastSync": last_sync,
                "metrics": {
                    "bookings": {
                        "total": len(ota_bookings),
                        "confirmed": len(confirmed_bookings),
                        "cancelled": len(cancelled_bookings),
                        "pending": len(pending_bookings),
                        "cancellationRate": (len(cancelled_bookings) / len(ota_bookings) * 100) if ota_bookings else 0.0,
                        "drillDownUrl": f"/api/v1/channel-manager/performance/{ota.ota_code}/bookings",
                        "clickable": True
                    },
                    "revenue": {
                        "gross": confirmed_revenue,
                        "cancelled": cancelled_revenue,
                        "net": net_revenue,
                        "afterCommission": net_revenue_after_commission,
                        "avgBookingValue": avg_booking_value,
                        "drillDownUrl": f"/api/v1/channel-manager/performance/{ota.ota_code}/revenue",
                        "clickable": True
                    },
                    "commission": {
                        "rate": commission_rate,
                        "amount": commission_amount
                    },
                    "rating": {
                        "value": round(avg_rating, 2),
                        "totalReviews": len(ota_reviews),
                        "drillDownUrl": f"/api/v1/channel-manager/performance/{ota.ota_code}/reviews",
                        "clickable": True
                    },
                    "conversionRate": conversion_rate,
                    "trends": {
                        "revenue": revenue_trend,
                        "bookings": bookings_trend
                    }
                }
            }
            
            channel_performance.append(channel_data)
            total_bookings += len(confirmed_bookings)
            total_revenue += confirmed_revenue
            total_commission += commission_amount
        
        # Sort by revenue (descending)
        channel_performance.sort(key=lambda x: x["metrics"]["revenue"]["net"], reverse=True)
        
        # Calculate summary
        net_revenue_total = total_revenue - total_commission
        
        summary = {
            "totalChannels": len(channel_performance),
            "totalBookings": total_bookings,
            "totalRevenue": total_revenue,
            "totalCommission": total_commission,
            "netRevenue": net_revenue_total,
            "avgCommissionRate": (total_commission / total_revenue * 100) if total_revenue > 0 else 0.0
        }
        
        logger.info(f"[CHANNEL_MANAGER] Detailed performance - Channels: {len(channel_performance)}, Bookings: {total_bookings}, Revenue: ${total_revenue:.2f}")
        print(f"[CHANNEL_MANAGER] Detailed performance calculated for {len(channel_performance)} channel(s)")
        
        return {
            "success": True,
            "data": {
                "period": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat(),
                    "days": (end_date - start_date).days + 1
                },
                "channels": channel_performance,
                "summary": summary
            }
        }
    
    except Exception as e:
        error_msg = f"Error getting detailed OTA performance: {str(e)}"
        logger.error(f"[CHANNEL_MANAGER] {error_msg}", exc_info=True)
        raise HTTPException(status_code=500, detail=error_msg)


# ============== DRILL-DOWN ENDPOINTS ==============

@router.get("/performance/{ota_code}/bookings", tags=["Channel Manager"])
async def get_ota_bookings(
    ota_code: str = Path(..., description="OTA code (e.g., BOOKING, EXPEDIA)"),
    start_date: Optional[date] = Query(None, alias="startDate", description="Start date for filtering (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, alias="endDate", description="End date for filtering (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by booking status"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get detailed bookings list for a specific OTA - drill-down from performance metrics"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /performance/{ota_code}/bookings - User: {user_info}")
    
    # Get OTA connection
    ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
    ota_result = await session.execute(ota_stmt)
    ota = ota_result.scalar_one_or_none()
    
    if not ota:
        raise HTTPException(status_code=404, detail=f"OTA connection not found for code: {ota_code}")
    
    # Build query for bookings
    bookings_stmt = select(Booking).where(
        Booking.booking_source.like(f"%{ota_code.lower()}%")
    )
    
    # Apply date filters
    if start_date:
        bookings_stmt = bookings_stmt.where(
            Booking.created_at >= datetime.combine(start_date, datetime.min.time())
        )
    if end_date:
        bookings_stmt = bookings_stmt.where(
            Booking.created_at <= datetime.combine(end_date, datetime.max.time())
        )
    
    # Apply status filter
    if status:
        bookings_stmt = bookings_stmt.where(Booking.status == status)
    
    # Get total count
    count_stmt = select(func.count()).select_from(bookings_stmt.subquery())
    total_result = await session.execute(count_stmt)
    total = total_result.scalar_one()
    
    # Apply pagination
    bookings_stmt = bookings_stmt.order_by(Booking.created_at.desc())
    bookings_stmt = bookings_stmt.offset((page - 1) * page_size).limit(page_size)
    
    result = await session.execute(bookings_stmt)
    bookings = result.scalars().all()
    
    # Get guest information for each booking
    items = []
    for booking in bookings:
        guest = await session.get(Guest, booking.guest_id)
        room_type = await session.get(RoomType, booking.room_type_id) if booking.room_type_id else None
        
        items.append({
            "id": booking.id,
            "bookingNumber": booking.booking_number,
            "confirmationCode": booking.confirmation_code,
            "guest": {
                "id": guest.id if guest else None,
                "name": f"{guest.first_name} {guest.last_name}" if guest else "Unknown",
                "email": guest.email if guest else None,
                "phone": guest.phone if guest else None
            },
            "roomType": {
                "id": room_type.id if room_type else None,
                "name": room_type.name if room_type else "Unknown"
            },
            "arrivalDate": booking.arrival_date.isoformat() if booking.arrival_date else None,
            "departureDate": booking.departure_date.isoformat() if booking.departure_date else None,
            "nights": booking.nights,
            "status": booking.status,
            "paymentStatus": booking.payment_status,
            "basePrice": booking.base_price,
            "taxes": booking.taxes,
            "serviceFee": booking.service_fee,
            "totalPrice": booking.total_price,
            "commissionRate": booking.commission_rate,
            "commissionAmount": booking.commission_amount,
            "netRevenue": booking.net_revenue,
            "createdAt": booking.created_at.isoformat() if booking.created_at else None,
            "checkInDate": booking.check_in_date.isoformat() if booking.check_in_date else None,
            "checkOutDate": booking.check_out_date.isoformat() if booking.check_out_date else None
        })
    
    return {
        "success": True,
        "data": {
            "ota": {
                "id": ota.id,
                "name": ota.ota_name,
                "code": ota.ota_code
            },
            "items": items,
            "pagination": {
                "page": page,
                "pageSize": page_size,
                "total": total,
                "totalPages": (total + page_size - 1) // page_size
            },
            "summary": {
                "totalBookings": total,
                "totalRevenue": sum(b.total_price for b in bookings if b.status != "cancelled"),
                "totalNetRevenue": sum(b.net_revenue or b.total_price for b in bookings if b.status != "cancelled"),
                "avgBookingValue": sum(b.total_price for b in bookings if b.status != "cancelled") / len([b for b in bookings if b.status != "cancelled"]) if bookings else 0.0
            }
        }
    }


@router.get("/performance/{ota_code}/reviews", tags=["Channel Manager"])
async def get_ota_reviews(
    ota_code: str = Path(..., description="OTA code (e.g., BOOKING, EXPEDIA)"),
    start_date: Optional[date] = Query(None, alias="startDate", description="Start date for filtering (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, alias="endDate", description="End date for filtering (YYYY-MM-DD)"),
    min_rating: Optional[float] = Query(None, ge=1.0, le=5.0, description="Minimum rating filter"),
    max_rating: Optional[float] = Query(None, ge=1.0, le=5.0, description="Maximum rating filter"),
    sentiment: Optional[str] = Query(None, description="Filter by sentiment (positive, neutral, negative)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get detailed reviews/ratings for a specific OTA - drill-down from performance metrics"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /performance/{ota_code}/reviews - User: {user_info}")
    
    # Get OTA connection
    ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
    ota_result = await session.execute(ota_stmt)
    ota = ota_result.scalar_one_or_none()
    
    if not ota:
        raise HTTPException(status_code=404, detail=f"OTA connection not found for code: {ota_code}")
    
    # Map OTA code to review source (e.g., BOOKING -> booking_com, EXPEDIA -> expedia)
    source_map = {
        "BOOKING": "booking_com",
        "EXPEDIA": "expedia",
        "AGODA": "agoda",
        "AIRBNB": "airbnb",
        "TRIPADVISOR": "tripadvisor"
    }
    review_source = source_map.get(ota_code.upper(), ota_code.lower())
    
    # Build query for reviews
    reviews_stmt = select(Review).where(
        or_(
            Review.source == review_source,
            Review.source.like(f"%{ota_code.lower()}%"),
            Review.source.like(f"%{ota.ota_name.lower()}%")
        )
    )
    
    # Apply date filters
    if start_date:
        reviews_stmt = reviews_stmt.where(
            Review.review_date >= datetime.combine(start_date, datetime.min.time())
        )
    if end_date:
        reviews_stmt = reviews_stmt.where(
            Review.review_date <= datetime.combine(end_date, datetime.max.time())
        )
    
    # Apply rating filters
    if min_rating:
        reviews_stmt = reviews_stmt.where(Review.overall_rating >= min_rating)
    if max_rating:
        reviews_stmt = reviews_stmt.where(Review.overall_rating <= max_rating)
    
    # Apply sentiment filter
    if sentiment:
        reviews_stmt = reviews_stmt.where(Review.sentiment == sentiment)
    
    # Get total count
    count_stmt = select(func.count()).select_from(reviews_stmt.subquery())
    total_result = await session.execute(count_stmt)
    total = total_result.scalar_one()
    
    # Apply pagination
    reviews_stmt = reviews_stmt.order_by(Review.review_date.desc())
    reviews_stmt = reviews_stmt.offset((page - 1) * page_size).limit(page_size)
    
    result = await session.execute(reviews_stmt)
    reviews = result.scalars().all()
    
    # Get guest information for each review
    items = []
    rating_breakdown = {"5": 0, "4": 0, "3": 0, "2": 0, "1": 0}
    sentiment_breakdown = {"positive": 0, "neutral": 0, "negative": 0}
    
    for review in reviews:
        guest = await session.get(Guest, review.guest_id) if review.guest_id else None
        booking = await session.get(Booking, review.booking_id) if review.booking_id else None
        
        # Update breakdowns
        rating_key = str(int(review.overall_rating))
        if rating_key in rating_breakdown:
            rating_breakdown[rating_key] += 1
        
        if review.sentiment and review.sentiment in sentiment_breakdown:
            sentiment_breakdown[review.sentiment] += 1
        
        items.append({
            "id": review.id,
            "guest": {
                "id": guest.id if guest else None,
                "name": f"{guest.first_name} {guest.last_name}" if guest else "Anonymous",
                "email": guest.email if guest else None
            },
            "booking": {
                "id": booking.id if booking else None,
                "bookingNumber": booking.booking_number if booking else None
            },
            "overallRating": review.overall_rating,
            "ratings": {
                "cleanliness": review.cleanliness_rating,
                "service": review.service_rating,
                "location": review.location_rating,
                "value": review.value_rating,
                "amenities": review.amenities_rating
            },
            "title": review.title,
            "comment": review.comment,
            "pros": review.pros,
            "cons": review.cons,
            "sentiment": review.sentiment,
            "isVerified": review.is_verified,
            "isPublic": review.is_public,
            "helpfulCount": review.helpful_count,
            "reviewDate": review.review_date.isoformat() if review.review_date else None,
            "response": review.response,
            "respondedAt": review.responded_at.isoformat() if review.responded_at else None,
            "source": review.source
        })
    
    # Calculate average ratings
    avg_ratings = {}
    if reviews:
        avg_ratings = {
            "overall": sum(r.overall_rating for r in reviews) / len(reviews),
            "cleanliness": sum(r.cleanliness_rating for r in reviews if r.cleanliness_rating) / len([r for r in reviews if r.cleanliness_rating]) if any(r.cleanliness_rating for r in reviews) else None,
            "service": sum(r.service_rating for r in reviews if r.service_rating) / len([r for r in reviews if r.service_rating]) if any(r.service_rating for r in reviews) else None,
            "location": sum(r.location_rating for r in reviews if r.location_rating) / len([r for r in reviews if r.location_rating]) if any(r.location_rating for r in reviews) else None,
            "value": sum(r.value_rating for r in reviews if r.value_rating) / len([r for r in reviews if r.value_rating]) if any(r.value_rating for r in reviews) else None,
            "amenities": sum(r.amenities_rating for r in reviews if r.amenities_rating) / len([r for r in reviews if r.amenities_rating]) if any(r.amenities_rating for r in reviews) else None
        }
    
    return {
        "success": True,
        "data": {
            "ota": {
                "id": ota.id,
                "name": ota.ota_name,
                "code": ota.ota_code
            },
            "items": items,
            "pagination": {
                "page": page,
                "pageSize": page_size,
                "total": total,
                "totalPages": (total + page_size - 1) // page_size
            },
            "summary": {
                "totalReviews": total,
                "averageRatings": avg_ratings,
                "ratingBreakdown": rating_breakdown,
                "sentimentBreakdown": sentiment_breakdown,
                "verifiedReviews": sum(1 for r in reviews if r.is_verified),
                "reviewsWithResponse": sum(1 for r in reviews if r.response)
            }
        }
    }


@router.get("/performance/{ota_code}/revenue", tags=["Channel Manager"])
async def get_ota_revenue(
    ota_code: str = Path(..., description="OTA code (e.g., BOOKING, EXPEDIA)"),
    start_date: Optional[date] = Query(None, alias="startDate", description="Start date for filtering (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, alias="endDate", description="End date for filtering (YYYY-MM-DD)"),
    group_by: Optional[str] = Query("day", description="Group by: day, week, month"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get detailed revenue breakdown for a specific OTA - drill-down from performance metrics"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /performance/{ota_code}/revenue - User: {user_info}")
    
    # Get OTA connection
    ota_stmt = select(OTAConnection).where(OTAConnection.ota_code == ota_code)
    ota_result = await session.execute(ota_stmt)
    ota = ota_result.scalar_one_or_none()
    
    if not ota:
        raise HTTPException(status_code=404, detail=f"OTA connection not found for code: {ota_code}")
    
    # Set default date range (last 30 days if not specified)
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    # Get bookings in date range
    bookings_stmt = select(Booking).where(
        and_(
            Booking.booking_source.like(f"%{ota_code.lower()}%"),
            Booking.created_at >= datetime.combine(start_date, datetime.min.time()),
            Booking.created_at <= datetime.combine(end_date, datetime.max.time())
        )
    )
    bookings_result = await session.execute(bookings_stmt)
    bookings = bookings_result.scalars().all()
    
    # Group revenue by time period
    revenue_by_period = {}
    commission_rate = ota.commission_rate or 0.0
    
    for booking in bookings:
        if booking.status == "cancelled":
            continue
        
        # Determine period key based on group_by
        booking_date = booking.created_at.date() if booking.created_at else booking.arrival_date
        
        if group_by == "week":
            # Get week start (Monday)
            days_since_monday = booking_date.weekday()
            period_start = booking_date - timedelta(days=days_since_monday)
            period_key = period_start.isoformat()
        elif group_by == "month":
            period_key = booking_date.strftime("%Y-%m")
        else:  # day
            period_key = booking_date.isoformat()
        
        if period_key not in revenue_by_period:
            revenue_by_period[period_key] = {
                "period": period_key,
                "grossRevenue": 0.0,
                "commission": 0.0,
                "netRevenue": 0.0,
                "bookings": 0,
                "avgBookingValue": 0.0
            }
        
        revenue_by_period[period_key]["grossRevenue"] += booking.total_price
        commission_amount = (booking.total_price * commission_rate) / 100
        revenue_by_period[period_key]["commission"] += commission_amount
        revenue_by_period[period_key]["netRevenue"] += booking.net_revenue or (booking.total_price - commission_amount)
        revenue_by_period[period_key]["bookings"] += 1
    
    # Calculate averages and sort
    revenue_breakdown = []
    for period_key in sorted(revenue_by_period.keys()):
        period_data = revenue_by_period[period_key]
        if period_data["bookings"] > 0:
            period_data["avgBookingValue"] = period_data["grossRevenue"] / period_data["bookings"]
        revenue_breakdown.append(period_data)
    
    # Calculate totals
    total_gross = sum(b.total_price for b in bookings if b.status != "cancelled")
    total_commission = (total_gross * commission_rate) / 100
    total_net = sum(b.net_revenue or (b.total_price - (b.total_price * commission_rate) / 100) for b in bookings if b.status != "cancelled")
    
    return {
        "success": True,
        "data": {
            "ota": {
                "id": ota.id,
                "name": ota.ota_name,
                "code": ota.ota_code,
                "commissionRate": commission_rate
            },
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
                "groupBy": group_by
            },
            "breakdown": revenue_breakdown,
            "summary": {
                "totalBookings": len([b for b in bookings if b.status != "cancelled"]),
                "totalGrossRevenue": total_gross,
                "totalCommission": total_commission,
                "totalNetRevenue": total_net,
                "avgBookingValue": total_gross / len([b for b in bookings if b.status != "cancelled"]) if bookings else 0.0,
                "commissionRate": commission_rate
            }
        }
    }


# ============== BOOKINGS ENDPOINT (for Dashboard) ==============

@router.get("/bookings", tags=["Channel Manager"])
async def get_bookings_by_source(
    source: Optional[str] = Query(None, description="Booking source (e.g., CRS for dummy CM, Booking.com for Booking.com)"),
    limit: int = Query(50, ge=1, le=1000, description="Number of bookings to return"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get bookings filtered by source - used by Dashboard for OTA performance details"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /bookings - Source: {source}, Limit: {limit}, User: {user_info}")
    
    stmt = select(Booking)
    
    if source:
        stmt = stmt.where(Booking.booking_source == source)
    
    stmt = stmt.where(Booking.status != "cancelled")
    stmt = stmt.order_by(Booking.created_at.desc()).limit(limit)
    
    result = await session.execute(stmt)
    bookings = result.scalars().all()
    
    items = []
    for booking in bookings:
        guest = await session.get(Guest, booking.guest_id) if booking.guest_id else None
        room_type = await session.get(RoomType, booking.room_type_id) if booking.room_type_id else None
        
        guest_name = ""
        if guest:
            guest_name = f"{guest.first_name} {guest.last_name}".strip() or guest.email or "Guest"
        
        items.append({
            "id": booking.id,
            "guest": guest_name,
            "guestName": guest_name,
            "email": guest.email if guest else "",
            "phone": guest.phone if guest else "",
            "checkIn": booking.arrival_date.isoformat() if booking.arrival_date else None,
            "checkOut": booking.departure_date.isoformat() if booking.departure_date else None,
            "roomType": room_type.name if room_type else "",
            "amount": booking.total_price,
            "total": booking.total_price,
            "source": booking.booking_source or ""
        })
    
    return {
        "success": True,
        "data": {
            "items": items,
            "total": len(items)
        }
    }


@router.get("/stats/insights", tags=["Channel Manager"])
async def get_channel_insights(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get AI insights and recommendations"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /stats/insights - User: {user_info}")
    print(f"[CHANNEL_MANAGER] Getting channel insights")
    
    insights = []
    
    # Get OTA connections
    ota_stmt = select(OTAConnection)
    ota_result = await session.execute(ota_stmt)
    otas = ota_result.scalars().all()
    
    # Check for disconnected OTAs
    disconnected_otas = [ota for ota in otas if ota.connection_status == "disconnected"]
    if disconnected_otas:
        insights.append({
            "type": "warning",
            "title": "Disconnected OTAs",
            "message": f"{len(disconnected_otas)} OTA(s) are disconnected",
            "action": "Review and reconnect OTAs"
        })
    
    # Check for unmapped room types
    room_types_stmt = select(RoomType)
    room_types_result = await session.execute(room_types_stmt)
    room_types = room_types_result.scalars().all()
    
    mappings_stmt = select(OTARoomMapping).where(OTARoomMapping.is_active == True)
    mappings_result = await session.execute(mappings_stmt)
    mappings = mappings_result.scalars().all()
    mapped_room_type_ids = set(m.room_type_id for m in mappings)
    unmapped_room_types = [rt for rt in room_types if rt.id not in mapped_room_type_ids]
    
    if unmapped_room_types:
        insights.append({
            "type": "info",
            "title": "Unmapped Room Types",
            "message": f"{len(unmapped_room_types)} room type(s) are not mapped to any OTA",
            "action": "Map room types to OTAs for better distribution"
        })
    
    # Check for rate parity issues
    today = date.today()
    rate_parity_stmt = select(DailyRate).where(DailyRate.date == today)
    rate_parity_result = await session.execute(rate_parity_stmt)
    daily_rates = rate_parity_result.scalars().all()
    
    if daily_rates:
        # Group by room type
        room_type_rates = {}
        for dr in daily_rates:
            if dr.room_type_id not in room_type_rates:
                room_type_rates[dr.room_type_id] = []
            rate = dr.override_rate or dr.base_rate
            if rate:
                room_type_rates[dr.room_type_id].append(rate)

        # Build lookup from already-fetched room_types
        rt_lookup = {rt.id: rt for rt in room_types}

        for room_type_id, rates in room_type_rates.items():
            if len(rates) > 1:
                min_rate = min(rates)
                max_rate = max(rates)
                if min_rate > 0:
                    difference = ((max_rate - min_rate) / min_rate) * 100
                    if difference > 10:  # 10% threshold
                        rt = rt_lookup.get(room_type_id)
                        insights.append({
                            "type": "warning",
                            "title": "Rate Parity Issue",
                            "message": f"{rt.name if rt else 'Room type'} has {difference:.1f}% rate variance",
                            "action": "Review and align rates across channels"
                        })
    
    # Check for connection errors
    error_otas = [ota for ota in otas if ota.connection_status == "error"]
    if error_otas:
        # List which specific OTAs have errors
        error_ota_names = [ota.ota_name for ota in error_otas]
        if len(error_ota_names) == 1:
            error_message = f"{error_ota_names[0]} has a connection error"
        else:
            error_message = f"{len(error_otas)} OTAs have connection errors: {', '.join(error_ota_names)}"
        
        insights.append({
            "type": "error",
            "title": "OTA Connection Errors",
            "message": error_message,
            "action": "Check OTA credentials and connection settings",
            "errorOTAs": [
                {
                    "id": ota.id,
                    "name": ota.ota_name,
                    "code": ota.ota_code,
                    "errorMessage": ota.error_message
                }
                for ota in error_otas
            ]
        })
    
    # Success message if everything is good
    if not insights:
        insights.append({
            "type": "success",
            "title": "All Systems Operational",
            "message": "All OTAs are connected and room types are mapped",
            "action": "Continue monitoring"
        })
    
    logger.info(f"[CHANNEL_MANAGER] Generated {len(insights)} insights")
    print(f"[CHANNEL_MANAGER] Generated {len(insights)} insights")
    
    return {
        "success": True,
        "data": {
            "insights": insights
        }
    }

# ============== PROMOTIONS ENDPOINTS ==============

@router.get("/promotions", tags=["Channel Manager"])
async def get_promotions(
    status: Optional[str] = Query(None, description="Filter by status: active, scheduled, expired, inactive, all"),
    ota_code: Optional[str] = Query(None),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get all channel promotions"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /promotions - Status: {status}, OTA: {ota_code}, User: {user_info}")
    print(f"[CHANNEL_MANAGER] Getting promotions - Status: {status}, OTA: {ota_code}")
    
    today = date.today()
    stmt = select(PromoCode)
    
    if status == "active":
        stmt = stmt.where(
            and_(
                PromoCode.is_active == True,
                PromoCode.valid_from <= today,
                PromoCode.valid_until >= today
            )
        )
    elif status == "scheduled":
        stmt = stmt.where(PromoCode.valid_from > today)
    elif status == "expired":
        stmt = stmt.where(PromoCode.valid_until < today)
    elif status == "inactive":
        stmt = stmt.where(PromoCode.is_active == False)
    
    result = await session.execute(stmt)
    promos = result.scalars().all()
    
    items = []
    for promo in promos:
        # Parse otaCodes and roomTypes from applicable_rate_plans JSON or use defaults
        import json
        ota_codes = ["ALL"]
        room_types = ["ALL"]
        booking_window = None
        
        # Try to parse metadata from applicable_rate_plans or use a separate JSON field
        # For now, we'll use a simple approach - store in applicable_rate_plans as JSON
        if promo.applicable_rate_plans:
            try:
                metadata = json.loads(promo.applicable_rate_plans)
                if isinstance(metadata, dict):
                    ota_codes = metadata.get("otaCodes", ["ALL"])
                    room_types = metadata.get("roomTypes", ["ALL"])
                    if "bookingWindow" in metadata:
                        booking_window = metadata["bookingWindow"]
            except:
                pass
        
        items.append({
            "id": promo.id,
            "name": promo.name,
            "description": promo.description,
            "code": promo.code,
            "discountType": promo.discount_type,
            "discountValue": promo.discount_value,
            "validFrom": promo.valid_from.isoformat(),
            "validTo": promo.valid_until.isoformat(),
            "minStay": promo.min_stay,
            "maxStay": promo.max_stay,
            "usageCount": promo.usage_count,
            "usageLimit": promo.usage_limit,
            "isActive": promo.is_active,
            "otaCodes": ota_codes,
            "roomTypes": room_types,
            "bookingWindow": booking_window,
            "createdAt": promo.created_at.isoformat() if promo.created_at else None,
            "updatedAt": promo.updated_at.isoformat() if promo.updated_at else None
        })
    
    logger.info(f"[CHANNEL_MANAGER] Found {len(items)} promotions")
    print(f"[CHANNEL_MANAGER] Found {len(items)} promotions")
    
    return {"success": True, "data": {"items": items, "total": len(items)}}


@router.get("/promotions/{promotion_id}", tags=["Channel Manager"])
async def get_promotion(
    promotion_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Get specific promotion"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] GET /promotions/{promotion_id} - User: {user_info}")
    print(f"[CHANNEL_MANAGER] Getting promotion: {promotion_id}")
    
    promo = await session.get(PromoCode, promotion_id)
    if not promo:
        raise HTTPException(status_code=404, detail="Promotion not found")
    
    # Parse otaCodes and roomTypes from applicable_rate_plans JSON
    import json
    ota_codes = ["ALL"]
    room_types = ["ALL"]
    booking_window = None
    
    if promo.applicable_rate_plans:
        try:
            metadata = json.loads(promo.applicable_rate_plans)
            if isinstance(metadata, dict):
                ota_codes = metadata.get("otaCodes", ["ALL"])
                room_types = metadata.get("roomTypes", ["ALL"])
                if "bookingWindow" in metadata:
                    booking_window = metadata["bookingWindow"]
        except:
            pass
    
    return {
        "success": True,
        "data": {
            "id": promo.id,
            "name": promo.name,
            "description": promo.description,
            "code": promo.code,
            "discountType": promo.discount_type,
            "discountValue": promo.discount_value,
            "validFrom": promo.valid_from.isoformat(),
            "validTo": promo.valid_until.isoformat(),
            "minStay": promo.min_stay,
            "maxStay": promo.max_stay,
            "usageCount": promo.usage_count,
            "usageLimit": promo.usage_limit,
            "isActive": promo.is_active,
            "otaCodes": ota_codes,
            "roomTypes": room_types,
            "bookingWindow": booking_window,
            "createdAt": promo.created_at.isoformat() if promo.created_at else None,
            "updatedAt": promo.updated_at.isoformat() if promo.updated_at else None
        }
    }


@router.post("/promotions", tags=["Channel Manager"])
async def create_promotion(
    payload: PromotionCreateRequest,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Create new promotion"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] POST /promotions - Name: {payload.name}, User: {user_info}")
    print(f"[CHANNEL_MANAGER] Creating promotion: {payload.name}")
    
    # Store otaCodes, roomTypes, and bookingWindow in applicable_rate_plans as JSON
    import json
    metadata = {
        "otaCodes": payload.otaCodes if hasattr(payload, 'otaCodes') else ["ALL"],
        "roomTypes": payload.roomTypes if hasattr(payload, 'roomTypes') else ["ALL"]
    }
    if hasattr(payload, 'bookingWindow') and payload.bookingWindow:
        metadata["bookingWindow"] = payload.bookingWindow
    
    # Generate unique promo code from name
    base_code = payload.name.upper().replace(" ", "_")
    code = base_code
    suffix = 1
    while True:
        existing = await session.execute(select(PromoCode).where(PromoCode.code == code))
        if not existing.scalars().first():
            break
        suffix += 1
        code = f"{base_code}_{suffix}"

    # Create promo code
    promo = PromoCode(
        code=code,
        name=payload.name,
        description=payload.description,
        discount_type=payload.discountType,
        discount_value=payload.discountValue,
        valid_from=datetime.fromisoformat(payload.validFrom).date(),
        valid_until=datetime.fromisoformat(payload.validTo).date(),
        min_stay=payload.minStay,
        max_stay=None,
        is_active=True,
        applicable_rate_plans=json.dumps(metadata)  # Store metadata as JSON
    )
    
    session.add(promo)
    await session.commit()
    await session.refresh(promo)
    
    logger.info(f"[CHANNEL_MANAGER] Created promotion: {promo.id}")
    print(f"[CHANNEL_MANAGER] Created promotion: {promo.id}")
    
    return {
        "success": True,
        "data": {
            "id": promo.id,
            "name": promo.name,
            "code": promo.code
        }
    }


@router.put("/promotions/{promotion_id}", tags=["Channel Manager"])
async def update_promotion(
    promotion_id: int,
    payload: Dict[str, Any] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Update promotion"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] PUT /promotions/{promotion_id} - User: {user_info}")
    print(f"[CHANNEL_MANAGER] Updating promotion: {promotion_id}")
    
    promo = await session.get(PromoCode, promotion_id)
    if not promo:
        raise HTTPException(status_code=404, detail="Promotion not found")
    
    if "name" in payload:
        promo.name = payload["name"]
    if "description" in payload:
        promo.description = payload["description"]
    if "discountType" in payload:
        promo.discount_type = payload["discountType"]
    if "discountValue" in payload:
        promo.discount_value = payload["discountValue"]
    if "validFrom" in payload:
        promo.valid_from = datetime.fromisoformat(payload["validFrom"]).date()
    if "validTo" in payload:
        promo.valid_until = datetime.fromisoformat(payload["validTo"]).date()
    if "minStay" in payload:
        promo.min_stay = payload["minStay"]
    
    # Update otaCodes, roomTypes, and bookingWindow if provided
    import json
    if "otaCodes" in payload or "roomTypes" in payload or "bookingWindow" in payload:
        metadata = {}
        if promo.applicable_rate_plans:
            try:
                metadata = json.loads(promo.applicable_rate_plans)
            except:
                pass
        
        if "otaCodes" in payload:
            metadata["otaCodes"] = payload["otaCodes"]
        if "roomTypes" in payload:
            metadata["roomTypes"] = payload["roomTypes"]
        if "bookingWindow" in payload:
            if payload["bookingWindow"]:
                metadata["bookingWindow"] = payload["bookingWindow"]
            elif "bookingWindow" in metadata:
                del metadata["bookingWindow"]
        
        promo.applicable_rate_plans = json.dumps(metadata)
    
    promo.updated_at = datetime.utcnow()
    await session.commit()
    
    logger.info(f"[CHANNEL_MANAGER] Updated promotion: {promotion_id}")
    print(f"[CHANNEL_MANAGER] Updated promotion: {promotion_id}")
    
    return {"success": True, "message": "Promotion updated"}


@router.delete("/promotions/{promotion_id}", tags=["Channel Manager"])
async def delete_promotion(
    promotion_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Delete promotion"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] DELETE /promotions/{promotion_id} - User: {user_info}")
    print(f"[CHANNEL_MANAGER] Deleting promotion: {promotion_id}")
    
    promo = await session.get(PromoCode, promotion_id)
    if not promo:
        raise HTTPException(status_code=404, detail="Promotion not found")
    
    await session.delete(promo)
    await session.commit()
    
    logger.info(f"[CHANNEL_MANAGER] Deleted promotion: {promotion_id}")
    print(f"[CHANNEL_MANAGER] Deleted promotion: {promotion_id}")
    
    return {"success": True, "message": "Promotion deleted"}


@router.put("/promotions/{promotion_id}/toggle", tags=["Channel Manager"])
async def toggle_promotion(
    promotion_id: int,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Toggle promotion active status"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] PUT /promotions/{promotion_id}/toggle - User: {user_info}")
    print(f"[CHANNEL_MANAGER] Toggling promotion: {promotion_id}")
    
    promo = await session.get(PromoCode, promotion_id)
    if not promo:
        raise HTTPException(status_code=404, detail="Promotion not found")
    
    promo.is_active = not promo.is_active
    promo.updated_at = datetime.utcnow()
    await session.commit()
    
    logger.info(f"[CHANNEL_MANAGER] Toggled promotion {promotion_id} to {promo.is_active}")
    print(f"[CHANNEL_MANAGER] Toggled promotion {promotion_id} to {promo.is_active}")
    
    return {
        "success": True,
        "data": {
            "isActive": promo.is_active
        }
    }


@router.post("/promotions/{promotion_id}/apply", tags=["Channel Manager"])
async def apply_promotion(
    promotion_id: int,
    payload: Dict[str, Any] = Body(...),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: Optional[User] = Depends(get_current_user_optional)
):
    """Apply promotion to specific OTAs"""
    user_info = current_user.email if current_user else "internal_service"
    logger.info(f"[CHANNEL_MANAGER] POST /promotions/{promotion_id}/apply - User: {user_info}")
    print(f"[CHANNEL_MANAGER] Applying promotion {promotion_id} to OTAs")
    
    promo = await session.get(PromoCode, promotion_id)
    if not promo:
        raise HTTPException(status_code=404, detail="Promotion not found")
    
    ota_codes = payload.get("otaCodes", [])
    
    logger.info(f"[CHANNEL_MANAGER] Applying promotion to OTAs: {ota_codes}")
    print(f"[CHANNEL_MANAGER] Applying promotion to OTAs: {ota_codes}")
    
    # In a real implementation, this would update PromotionApplicability records
    # For now, just return success
    
    return {"success": True, "message": f"Promotion applied to {len(ota_codes)} OTA(s)"}
