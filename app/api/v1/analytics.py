"""
Advanced Analytics API Endpoints
Provides AI-powered analytics, natural language BI, and predictive insights
"""
from datetime import date
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, Query, HTTPException, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlmodel.ext.asyncio.session import AsyncSession
import json
import csv
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# Excel export support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.db.session import get_tenant_session
from app.api.v1.auth import get_current_user
from app.models.user import User
from app.services.advanced_analytics_service import AdvancedAnalyticsService

router = APIRouter()


# ==================== Request/Response Models ====================

class NaturalLanguageQuery(BaseModel):
    """Natural language analytics query"""
    query: str = Field(..., description="The analytics question in plain English")
    session_id: Optional[str] = Field(None, description="Session ID for context retention")


class ScenarioRequest(BaseModel):
    """Scenario analysis request"""
    type: str = Field(..., description="Scenario type: rate_change, staffing_change")
    change_percent: float = Field(..., description="Percentage change to model")
    additional_params: Optional[Dict[str, Any]] = Field(None, description="Additional scenario parameters")


class ReportRequest(BaseModel):
    """Report generation request"""
    report_type: str = Field(..., description="Type of report to generate")
    date_range: Optional[str] = Field("this_month", description="Date range for the report")
    export_format: Optional[str] = Field(None, description="Export format: pdf, excel, csv")
    custom_metrics: Optional[List[str]] = Field(None, description="Custom metrics to include")
    grouping: Optional[str] = Field(None, description="Data grouping: daily, weekly, monthly")


class AlertThreshold(BaseModel):
    """Alert threshold configuration"""
    metric: str
    threshold_value: float
    condition: str = Field(..., description="Condition: above, below, change_percent")
    severity: str = Field("warning", description="Alert severity: info, warning, critical")


class DashboardWidget(BaseModel):
    """Dashboard widget configuration"""
    id: str
    type: str = Field(..., description="Widget type: kpi, chart, table, gauge")
    metric: str
    title: str
    position: Dict[str, int] = Field(..., description="Grid position: x, y, w, h")
    config: Optional[Dict[str, Any]] = None


class DashboardLayout(BaseModel):
    """Custom dashboard layout"""
    name: str
    widgets: List[DashboardWidget]


# ==================== KPI Dashboard Endpoints ====================

@router.get("/dashboard")
async def get_analytics_dashboard(
    date_range: str = Query("today", description="Date range: today, week, month, year, ytd, last_7_days, last_30_days, etc."),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """
    Get comprehensive KPI dashboard with AI insights

    Returns all major KPIs, trend analysis, AI-generated insights, and detected anomalies.

    Supported date_range values:
    - today, yesterday
    - week, this_week, last_week
    - month, this_month, last_month
    - year, ytd, last_year
    - last_7_days, last_30_days, last_90_days, last_365_days
    """
    service = AdvancedAnalyticsService(session)
    try:
        return await service.get_kpi_dashboard(date_range)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/dashboard/kpis")
async def get_kpis(
    metrics: str = Query("revenue,occupancy,bookings,adr,revpar", description="Comma-separated metrics"),
    date_range: str = Query("today", description="Date range: today, week, month, year, ytd, last_30_days, etc."),
    compare: bool = Query(False, description="Include comparison with previous period"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get specific KPIs with optional comparison"""
    service = AdvancedAnalyticsService(session)
    try:
        dashboard = await service.get_kpi_dashboard(date_range)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    requested_metrics = [m.strip() for m in metrics.split(",")]
    filtered_kpis = {k: v for k, v in dashboard["kpis"].items() if k in requested_metrics}

    return {
        "period": dashboard["period"],
        "kpis": filtered_kpis,
        "generated_at": dashboard["generated_at"]
    }


@router.get("/dashboard/realtime")
async def get_realtime_metrics(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get real-time operational metrics"""
    service = AdvancedAnalyticsService(session)

    # Get today's data for real-time view
    dashboard = await service.get_kpi_dashboard("today")

    return {
        "timestamp": dashboard["generated_at"],
        "occupancy": dashboard["kpis"].get("occupancy", {}).get("value", 0),
        "arrivals_today": dashboard["kpis"].get("bookings", {}).get("value", 0),
        "revenue_today": dashboard["kpis"].get("revenue", {}).get("value", 0),
        "pending_tasks": 0,  # Will be populated from housekeeping/maintenance
        "active_alerts": len(dashboard.get("anomalies", []))
    }


# ==================== Natural Language BI Endpoints ====================

@router.post("/query")
async def query_analytics(
    request: NaturalLanguageQuery,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """
    Query analytics using natural language

    Ask questions like:
    - "What was our revenue last month?"
    - "Show me occupancy trends for this week"
    - "Compare bookings this month vs last month"
    - "Which room types have the highest ADR?"
    """
    service = AdvancedAnalyticsService(session)
    return await service.query_natural_language(
        request.query,
        request.session_id,
        current_user.id
    )


@router.get("/query/suggestions")
async def get_query_suggestions(
    context: Optional[str] = Query(None, description="Current page or context"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get suggested analytics queries based on context"""
    suggestions = {
        "general": [
            "What's our occupancy rate today?",
            "Show revenue for this month",
            "How many bookings do we have this week?",
            "Compare this month's performance to last month"
        ],
        "revenue": [
            "What's the ADR trend for this month?",
            "Revenue breakdown by room type",
            "Compare RevPAR with last month",
            "Show daily revenue for the past week"
        ],
        "operations": [
            "How many arrivals are expected today?",
            "What's the housekeeping completion rate?",
            "Show pending maintenance requests",
            "Staff efficiency metrics"
        ],
        "guests": [
            "How many returning guests this month?",
            "Guest satisfaction trends",
            "VIP guest statistics",
            "Guest demographics breakdown"
        ]
    }

    return {
        "suggestions": suggestions.get(context, suggestions["general"]),
        "popular": [
            "Today's revenue",
            "Occupancy rate",
            "Booking forecast"
        ]
    }


@router.get("/query/history")
async def get_query_history(
    session_id: Optional[str] = Query(None),
    limit: int = Query(20, le=100),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get analytics query history for the user"""
    # In a real implementation, this would fetch from database
    return {
        "queries": [],
        "session_id": session_id
    }


# ==================== Predictive Analytics Endpoints ====================

@router.get("/predictions/{metric}")
async def get_predictions(
    metric: str,
    days_ahead: int = Query(30, le=90, description="Number of days to forecast"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """
    Get predictive forecasts for a metric

    Supported metrics: revenue, occupancy, bookings, adr, revpar
    """
    valid_metrics = ["revenue", "occupancy", "bookings", "adr", "revpar"]
    if metric not in valid_metrics:
        raise HTTPException(status_code=400, detail=f"Invalid metric. Choose from: {valid_metrics}")

    service = AdvancedAnalyticsService(session)
    predictions = await service.get_predictions(metric, days_ahead)

    return {
        "metric": metric,
        "forecast_period": f"{days_ahead} days",
        "predictions": predictions,
        "generated_at": predictions[0]["date"] if predictions else None
    }


@router.get("/predictions/demand")
async def get_demand_forecast(
    days_ahead: int = Query(90, le=180, description="Forecast horizon in days"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """
    Get comprehensive demand forecast

    Returns occupancy and revenue predictions with high/low demand period identification.
    """
    service = AdvancedAnalyticsService(session)
    return await service.get_demand_forecast(days_ahead)


@router.post("/predictions/scenario")
async def run_scenario_analysis(
    request: ScenarioRequest,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """
    Run what-if scenario analysis

    Model the impact of changes like:
    - Rate changes (price elasticity modeling)
    - Staffing level changes
    - Marketing spend changes
    """
    valid_types = ["rate_change", "staffing_change", "marketing_change"]
    if request.type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid scenario type. Choose from: {valid_types}")

    service = AdvancedAnalyticsService(session)
    return await service.run_scenario_analysis({
        "type": request.type,
        "change_percent": request.change_percent,
        **(request.additional_params or {})
    })


# ==================== Anomaly Detection Endpoints ====================

@router.get("/anomalies")
async def detect_anomalies(
    lookback_days: int = Query(30, le=90, description="Days to analyze for anomalies"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """
    Detect anomalies in key metrics

    Identifies unusual patterns in revenue, occupancy, and bookings
    using statistical analysis.
    """
    service = AdvancedAnalyticsService(session)
    anomalies = await service.detect_anomalies(lookback_days)

    return {
        "lookback_period": f"{lookback_days} days",
        "anomalies": anomalies,
        "total_detected": len(anomalies),
        "critical_count": len([a for a in anomalies if a.get("severity") == "critical"]),
        "warning_count": len([a for a in anomalies if a.get("severity") == "warning"])
    }


@router.get("/insights")
async def get_ai_insights(
    date_range: str = Query("this_week", description="Date range: today, week, month, year, ytd, last_30_days, etc."),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get AI-highlighted insights and opportunities"""
    service = AdvancedAnalyticsService(session)
    try:
        dashboard = await service.get_kpi_dashboard(date_range)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "period": dashboard["period"],
        "insights": dashboard["insights"],
        "anomalies": dashboard["anomalies"],
        "opportunities": [
            i for i in dashboard["insights"]
            if i.get("trend") == "up" or "opportunity" in i.get("title", "").lower()
        ]
    }


# ==================== AI Report Builder Endpoints ====================

@router.post("/reports/generate")
async def generate_report(
    request: ReportRequest,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """
    Generate AI-powered report

    Available report types:
    - daily_operations: Daily operational summary
    - revenue_analysis: Revenue performance analysis
    - housekeeping_efficiency: Housekeeping metrics and efficiency
    - guest_analytics: Guest insights and retention metrics
    - executive_summary: High-level KPI summary with outlook
    """
    valid_types = ["daily_operations", "revenue_analysis", "housekeeping_efficiency",
                   "guest_analytics", "executive_summary"]
    if request.report_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid report type. Choose from: {valid_types}")

    service = AdvancedAnalyticsService(session)
    report = await service.generate_report(
        request.report_type,
        {
            "date_range": request.date_range,
            "custom_metrics": request.custom_metrics,
            "grouping": request.grouping
        }
    )

    # Handle export if requested
    if request.export_format:
        return await export_report(report, request.export_format)

    return report


@router.get("/reports/templates")
async def get_report_templates(
    role: Optional[str] = Query(None, description="Filter by user role"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get available report templates with role-based suggestions"""
    templates = [
        {
            "id": "daily_operations",
            "name": "Daily Operations Report",
            "description": "Comprehensive daily operational summary including arrivals, departures, and tasks",
            "recommended_for": ["manager", "admin", "frontdesk"],
            "frequency": "daily"
        },
        {
            "id": "revenue_analysis",
            "name": "Revenue Analysis Report",
            "description": "Detailed revenue metrics, trends, and forecasts with ADR/RevPAR analysis",
            "recommended_for": ["manager", "admin", "finance"],
            "frequency": "weekly"
        },
        {
            "id": "housekeeping_efficiency",
            "name": "Housekeeping Efficiency Report",
            "description": "Task completion rates, staff performance, and room turnover metrics",
            "recommended_for": ["manager", "housekeeping"],
            "frequency": "daily"
        },
        {
            "id": "guest_analytics",
            "name": "Guest Analytics Report",
            "description": "Guest demographics, retention rates, and satisfaction insights",
            "recommended_for": ["manager", "admin", "frontdesk"],
            "frequency": "monthly"
        },
        {
            "id": "executive_summary",
            "name": "Executive Summary",
            "description": "High-level KPI overview with AI-generated insights and 30-day outlook",
            "recommended_for": ["manager", "admin"],
            "frequency": "weekly"
        }
    ]

    if role:
        # Case-insensitive role matching
        role_lower = role.lower()
        templates = [t for t in templates if role_lower in t["recommended_for"]]

    return {
        "templates": templates,
        "suggested": templates[0] if templates else None
    }


@router.get("/reports/{report_id}/export")
async def export_report_by_id(
    report_id: str,
    format: str = Query("pdf", description="Export format: pdf, excel, csv"),
    date_range: str = Query("this_month"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Export a generated report in the specified format"""
    service = AdvancedAnalyticsService(session)
    report = await service.generate_report(report_id, {"date_range": date_range})
    return await export_report(report, format)


async def export_report(report: Dict[str, Any], format: str) -> Response:
    """Export report to specified format"""

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header info
        writer.writerow(["Report Type", report.get("report_type", "N/A")])
        writer.writerow(["Generated At", report.get("generated_at", "N/A")])
        writer.writerow([])

        # Write metrics
        if "metrics" in report:
            writer.writerow(["Metric", "Value"])
            for key, value in report["metrics"].items():
                if isinstance(value, dict):
                    writer.writerow([key, value.get("value", value)])
                else:
                    writer.writerow([key, value])

        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={report.get('report_type', 'report')}.csv"}
        )

    elif format == "json":
        return Response(
            content=json.dumps(report, indent=2, default=str),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={report.get('report_type', 'report')}.json"}
        )

    elif format == "pdf":
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"<b>{report.get('report_type', 'Report').replace('_', ' ').title()}</b>", styles['Title']))
        elements.append(Spacer(1, 12))

        # Period
        period = report.get("period", {})
        elements.append(Paragraph(f"Period: {period.get('start', 'N/A')} to {period.get('end', 'N/A')}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Metrics table
        if "metrics" in report:
            data = [["Metric", "Value"]]
            for key, value in report["metrics"].items():
                if isinstance(value, dict):
                    data.append([key.replace("_", " ").title(), str(value.get("value", value))])
                else:
                    data.append([key.replace("_", " ").title(), str(value)])

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={report.get('report_type', 'report')}.pdf"}
        )

    elif format == "excel" or format == "xlsx":
        if not EXCEL_AVAILABLE:
            raise HTTPException(status_code=500, detail="Excel export not available. Please install openpyxl.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Report Title
        ws['A1'] = report.get('report_type', 'Report').replace('_', ' ').title()
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')

        # Generated timestamp
        ws['A2'] = "Generated:"
        ws['B2'] = report.get('generated_at', 'N/A')

        # Period
        period = report.get("period", {})
        ws['A3'] = "Period:"
        ws['B3'] = f"{period.get('start', 'N/A')} to {period.get('end', 'N/A')}"

        # Blank row
        current_row = 5

        # Metrics section
        if "metrics" in report:
            # Header
            ws.cell(row=current_row, column=1, value="Metric").font = header_font
            ws.cell(row=current_row, column=1).fill = header_fill
            ws.cell(row=current_row, column=1).alignment = header_alignment
            ws.cell(row=current_row, column=1).border = thin_border

            ws.cell(row=current_row, column=2, value="Value").font = header_font
            ws.cell(row=current_row, column=2).fill = header_fill
            ws.cell(row=current_row, column=2).alignment = header_alignment
            ws.cell(row=current_row, column=2).border = thin_border

            current_row += 1

            # Data rows
            for key, value in report["metrics"].items():
                ws.cell(row=current_row, column=1, value=key.replace("_", " ").title())
                ws.cell(row=current_row, column=1).border = thin_border

                if isinstance(value, dict):
                    ws.cell(row=current_row, column=2, value=str(value.get("value", value)))
                else:
                    ws.cell(row=current_row, column=2, value=str(value))
                ws.cell(row=current_row, column=2).border = thin_border

                current_row += 1

        # If there's data tables in the report
        if "data" in report and isinstance(report["data"], list) and len(report["data"]) > 0:
            current_row += 2
            ws.cell(row=current_row, column=1, value="Detailed Data").font = Font(bold=True, size=12)
            current_row += 1

            # Get headers from first item
            first_item = report["data"][0]
            if isinstance(first_item, dict):
                headers = list(first_item.keys())

                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header.replace("_", " ").title())
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                current_row += 1

                # Write data rows
                for item in report["data"]:
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=str(item.get(header, "")))
                        cell.border = thin_border
                    current_row += 1

        # Adjust column widths
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                try:
                    if ws.cell(row=row, column=col).value:
                        max_length = max(max_length, len(str(ws.cell(row=row, column=col).value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={report.get('report_type', 'report')}.xlsx"}
        )

    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {format}. Supported: csv, json, pdf, excel")


# ==================== Dashboard Customization Endpoints ====================

@router.post("/dashboard/layout")
async def save_dashboard_layout(
    layout: DashboardLayout,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Save custom dashboard layout for user"""
    # In production, save to database
    return {
        "message": "Dashboard layout saved",
        "layout_name": layout.name,
        "widget_count": len(layout.widgets)
    }


@router.get("/dashboard/layouts")
async def get_dashboard_layouts(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get user's saved dashboard layouts"""
    # Default layouts
    return {
        "layouts": [
            {
                "id": "default",
                "name": "Default Dashboard",
                "widgets": [
                    {"id": "revenue", "type": "kpi", "metric": "revenue", "title": "Revenue", "position": {"x": 0, "y": 0, "w": 3, "h": 1}},
                    {"id": "occupancy", "type": "gauge", "metric": "occupancy", "title": "Occupancy", "position": {"x": 3, "y": 0, "w": 3, "h": 1}},
                    {"id": "bookings", "type": "kpi", "metric": "bookings", "title": "Bookings", "position": {"x": 6, "y": 0, "w": 3, "h": 1}},
                    {"id": "adr", "type": "kpi", "metric": "adr", "title": "ADR", "position": {"x": 9, "y": 0, "w": 3, "h": 1}},
                    {"id": "trend", "type": "chart", "metric": "revenue", "title": "Revenue Trend", "position": {"x": 0, "y": 1, "w": 6, "h": 2}},
                    {"id": "insights", "type": "table", "metric": "insights", "title": "AI Insights", "position": {"x": 6, "y": 1, "w": 6, "h": 2}}
                ]
            }
        ],
        "active": "default"
    }


# ==================== Alert Configuration Endpoints ====================

@router.post("/alerts/threshold")
async def configure_alert_threshold(
    threshold: AlertThreshold,
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Configure custom alert threshold"""
    return {
        "message": "Alert threshold configured",
        "threshold": {
            "metric": threshold.metric,
            "value": threshold.threshold_value,
            "condition": threshold.condition,
            "severity": threshold.severity
        }
    }


@router.get("/alerts/active")
async def get_active_alerts(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get currently active alerts"""
    service = AdvancedAnalyticsService(session)
    anomalies = await service.detect_anomalies(7)  # Last 7 days

    return {
        "alerts": [
            {
                "id": f"alert_{i}",
                "type": "anomaly",
                "metric": a["metric"],
                "message": a["description"],
                "severity": a["severity"],
                "detected_at": a["date"]
            }
            for i, a in enumerate(anomalies)
        ],
        "total": len(anomalies)
    }


# ==================== Data Export Endpoints ====================

@router.get("/export/data")
async def export_analytics_data(
    metrics: str = Query("revenue,occupancy,bookings", description="Metrics to export"),
    date_range: str = Query("last_30_days"),
    format: str = Query("csv", description="Export format: csv, json"),
    aggregation: str = Query("daily", description="Data aggregation: daily, weekly, monthly"),
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Export raw analytics data for external analysis"""
    service = AdvancedAnalyticsService(session)

    # Parse query
    parsed = await service._parse_query(
        f"Show {metrics} {aggregation} for {date_range.replace('_', ' ')}",
        {}
    )
    data = await service._execute_analytics_query(parsed)

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)

        # Write data
        writer.writerow(["Date Range", data.get("period", {}).get("start"), "to", data.get("period", {}).get("end")])
        writer.writerow([])

        for metric, values in data.items():
            if metric == "period":
                continue
            writer.writerow([metric.upper()])
            if isinstance(values, dict):
                for k, v in values.items():
                    writer.writerow([k, v])
            else:
                writer.writerow(["value", values])
            writer.writerow([])

        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=analytics_export.csv"}
        )

    return data


# ==================== Benchmarking Endpoints ====================

@router.get("/benchmarks")
async def get_benchmarks(
    session: AsyncSession = Depends(get_tenant_session),
    current_user: User = Depends(get_current_user)
):
    """Get industry benchmarks for comparison"""
    # Sample benchmark data - in production, this would come from industry data sources
    return {
        "region": "Asia Pacific",
        "property_type": "Boutique Hotel",
        "benchmarks": {
            "occupancy": {"industry_avg": 68.5, "top_performer": 85.0},
            "adr": {"industry_avg": 12500.0, "top_performer": 18500.0},
            "revpar": {"industry_avg": 8500.0, "top_performer": 15500.0},
            "guest_satisfaction": {"industry_avg": 4.2, "top_performer": 4.8}
        },
        "last_updated": "2024-01-01"
    }
